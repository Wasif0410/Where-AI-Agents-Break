"""
build_inputs.py — generates all input files under /root/data and buggy project
files under /root/project.  Run once during Docker image build.
"""

import os
import json
import csv
import textwrap
from datetime import date, datetime

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
DATA_DIR = "/root/data"
PROJECT_DIR = "/root/project"

os.makedirs(DATA_DIR, exist_ok=True)
os.makedirs(PROJECT_DIR, exist_ok=True)

# Reduced scope: 9 accounts covering every violation type and workbook traps.
KEPT_ACCOUNT_IDS = frozenset({
    "acct_001",  # clean pass
    "acct_002",  # usage_tier_miscalculated
    "acct_003",  # wrong_plan
    "acct_004",  # pass — workbook plan_override (header row, stale sheet)
    "acct_005",  # expired_exception
    "acct_006",  # wrong_price_floor (seat floor)
    "acct_007",  # missing_grandfathering + integer discount
    "acct_008",  # pass — workbook price formulas + superseded row
    "acct_013",  # wrong_price_floor — formula override precedence
})


def _keep_account(row, id_col=0):
    return row[id_col] in KEPT_ACCOUNT_IDS


# ---------------------------------------------------------------------------
# 1. legacy_accounts.csv
# ---------------------------------------------------------------------------
_ACCOUNTS_ALL = [
    ("acct_001", "Northstar Labs",      "Legacy Starter",    "SMB",         "NA", 8,  "2026-04-01"),
    ("acct_002", "Bluefield Systems",   "Legacy Growth",     "Mid-Market",  "EU", 15, "2026-04-01"),
    ("acct_003", "Copperline AI",       "Legacy Pro",        "Enterprise",  "NA", 30, "2026-04-01"),
    ("acct_004", "Dawnridge Corp",      "Legacy Starter",    "SMB",         "NA", 5,  "2026-04-01"),
    ("acct_005", "Ember Analytics",     "Legacy Growth",     "Mid-Market",  "EU", 20, "2026-04-01"),
    ("acct_006", "Foxgate Solutions",   "Legacy Starter",    "SMB",         "NA", 25, "2026-04-01"),
    ("acct_007", "Greenfield SaaS",     "Legacy Growth",     "Mid-Market",  "EU", 35, "2026-04-01"),
    ("acct_008", "Highpoint Data",      "Legacy Pro",        "Enterprise",  "NA", 40, "2026-04-01"),
    ("acct_009", "Ironwood Systems",    "Legacy Growth",     "Mid-Market",  "EU", 22, "2026-04-01"),
    ("acct_010", "Juniper Cloud",       "Legacy Pro",        "Enterprise",  "NA", 60, "2026-04-01"),
    ("acct_011", "Keystone Metrics",    "Legacy Growth",     "Enterprise",  "NA", 28, "2026-04-01"),
    ("acct_012", "Landmark Tech",       "Legacy Enterprise", "Enterprise",  "EU", 50, "2026-04-01"),
    ("acct_013", "Meridian Cloud",      "Legacy Growth",     "Enterprise",  "EU", 55, "2026-04-01"),
    ("acct_014", "Redwood Analytics",   "Legacy Starter",    "Mid-Market", "EU", 40, "2026-04-01"),
    ("acct_015", "Nova Analytics",      "Legacy Pro",        "Enterprise",  "NA", 38, "2026-04-01"),
    ("acct_016", "Orion Systems",       "Legacy Growth",     "Enterprise",  "NA", 42, "2026-04-01"),
]
ACCOUNTS = [r for r in _ACCOUNTS_ALL if _keep_account(r)]

with open(os.path.join(DATA_DIR, "legacy_accounts.csv"), "w", newline="") as f:
    w = csv.writer(f)
    w.writerow(["account_id", "company_name", "legacy_plan", "segment", "region",
                "current_seats", "migration_effective_date"])
    w.writerows(ACCOUNTS)

# ---------------------------------------------------------------------------
# 2. product_usage_events.csv
# Trailing 90-day window for migration 2026-04-01 = 2026-01-01 to 2026-03-31.
# April events are included for ALL accounts so the "use migration month only"
# bug has equal data for each account.  For acct_002 April=High; acct_011
# April=Low — acct_002 is the usage-tier trap account in the reduced set.
# ---------------------------------------------------------------------------
_USAGE_EVENTS_ALL = [
    # acct_001: trailing 150,000 (Medium) | April 150,000 (Medium – same tier)
    ("acct_001", "2026-01-10", 60000,  8),
    ("acct_001", "2026-02-14", 50000,  7),
    ("acct_001", "2026-03-20", 40000,  8),
    ("acct_001", "2026-04-12", 150000, 8),   # April – same tier

    # acct_002: trailing 80,000 (Low) | April 650,000 (High – bug gives wrong tier)
    ("acct_002", "2026-01-10", 30000,  12),
    ("acct_002", "2026-02-18", 25000,  14),
    ("acct_002", "2026-03-22", 25000,  13),
    ("acct_002", "2026-04-05", 350000, 15),  # April – High
    ("acct_002", "2026-04-20", 300000, 15),  # April – High

    # acct_003: trailing 200,000 (Medium) | April 180,000 (Medium – same tier)
    ("acct_003", "2026-01-08", 70000,  28),
    ("acct_003", "2026-02-12", 80000,  30),
    ("acct_003", "2026-03-19", 50000,  29),
    ("acct_003", "2026-04-10", 180000, 29),  # April – same tier

    # acct_004: trailing 50,000 (Low) | April 60,000 (Low – same tier)
    ("acct_004", "2026-01-20", 20000,  4),
    ("acct_004", "2026-02-25", 15000,  5),
    ("acct_004", "2026-03-28", 15000,  5),
    ("acct_004", "2026-04-15", 60000,  5),   # April – same tier

    # acct_005: trailing 300,000 (Medium) | April 280,000 (Medium – same tier)
    ("acct_005", "2026-01-05", 100000, 18),
    ("acct_005", "2026-02-10", 120000, 20),
    ("acct_005", "2026-03-15", 80000,  19),
    ("acct_005", "2026-04-08", 280000, 19),  # April – same tier

    # acct_006: trailing 250,000 (Medium) | April 260,000 (Medium – same tier)
    ("acct_006", "2026-01-12", 80000,  23),
    ("acct_006", "2026-02-16", 100000, 25),
    ("acct_006", "2026-03-21", 70000,  24),
    ("acct_006", "2026-04-11", 260000, 24),  # April – same tier

    # acct_007: trailing 400,000 (Medium) | April 420,000 (Medium – same tier)
    ("acct_007", "2026-01-07", 100000, 33),
    ("acct_007", "2026-01-28", 50000,  34),
    ("acct_007", "2026-02-11", 140000, 35),
    ("acct_007", "2026-03-18", 110000, 34),
    ("acct_007", "2026-04-09", 420000, 35),  # April – same tier

    # acct_008: trailing 100,000 (Medium) | April 110,000 (Medium – same tier)
    ("acct_008", "2026-01-14", 35000,  38),
    ("acct_008", "2026-02-19", 35000,  40),
    ("acct_008", "2026-03-24", 30000,  39),
    ("acct_008", "2026-04-14", 110000, 39),  # April – same tier

    # acct_009: trailing 200,000 (Medium) | April 190,000 (Medium – same tier)
    ("acct_009", "2026-01-09", 65000,  20),
    ("acct_009", "2026-02-14", 75000,  22),
    ("acct_009", "2026-03-22", 60000,  21),
    ("acct_009", "2026-04-16", 190000, 21),  # April – same tier

    # acct_010: trailing 600,000 (High) | April 550,000 (High – same tier)
    ("acct_010", "2026-01-06",  150000, 58),
    ("acct_010", "2026-01-27",  100000, 59),
    ("acct_010", "2026-02-13",  150000, 60),
    ("acct_010", "2026-03-10",  100000, 58),
    ("acct_010", "2026-03-25",  100000, 59),
    ("acct_010", "2026-04-17",  550000, 60),  # April – same tier

    # acct_011: trailing 550,000 (High) | April 45,000 (Low – bug gives wrong tier)
    ("acct_011", "2026-01-11", 150000, 26),
    ("acct_011", "2026-01-30", 50000,  27),
    ("acct_011", "2026-02-15", 200000, 28),
    ("acct_011", "2026-03-20", 150000, 27),
    ("acct_011", "2026-04-08", 45000,  24),  # April – Low

    # acct_012: trailing 520,000 (High) | April 500,000 (High – same tier)
    ("acct_012", "2026-01-04", 130000, 48),
    ("acct_012", "2026-01-25", 60000,  49),
    ("acct_012", "2026-02-09", 180000, 50),
    ("acct_012", "2026-03-05", 80000,  48),
    ("acct_012", "2026-03-24", 70000,  49),
    ("acct_012", "2026-04-13", 500000, 50),  # April – same tier

    # acct_013: trailing 550,000 (High) | April 80,000 (Low – bug gives wrong tier)
    # Also has active price_override + grandfathering — cross-file trap
    ("acct_013", "2026-01-08", 180000, 53),
    ("acct_013", "2026-01-30",  50000, 54),
    ("acct_013", "2026-02-15", 200000, 55),
    ("acct_013", "2026-03-20", 120000, 53),
    ("acct_013", "2026-04-09",  80000, 52),  # April – Low

    # acct_014: trailing 70,000 (Low) | April 70,000 (Low – same tier)
    # Has plan_override to Scale Plus + grandfathering — rule-order trap
    ("acct_014", "2026-01-15", 25000, 38),
    ("acct_014", "2026-02-20", 25000, 40),
    ("acct_014", "2026-03-18", 20000, 39),
    ("acct_014", "2026-04-11", 70000, 40),  # April – same tier

    # acct_015: trailing 200,000 (Medium) | April 180,000 (Medium – same tier)
    ("acct_015", "2026-01-08", 70000,  36),
    ("acct_015", "2026-02-12", 80000,  38),
    ("acct_015", "2026-03-19", 50000,  37),
    ("acct_015", "2026-04-10", 180000, 37),

    # acct_016: trailing 560,000 (High) | April 520,000 (High – same tier)
    ("acct_016", "2026-01-06", 130000, 40),
    ("acct_016", "2026-01-28",  60000, 41),
    ("acct_016", "2026-02-14", 180000, 42),
    ("acct_016", "2026-03-22", 190000, 41),
    ("acct_016", "2026-04-15", 520000, 42),
]
USAGE_EVENTS = [r for r in _USAGE_EVENTS_ALL if _keep_account(r)]

with open(os.path.join(DATA_DIR, "product_usage_events.csv"), "w", newline="") as f:
    w = csv.writer(f)
    w.writerow(["account_id", "event_date", "api_calls", "active_users"])
    w.writerows(USAGE_EVENTS)

# ---------------------------------------------------------------------------
# 3. contract_terms.jsonl
# acct_007: grandfathered_until 2026-06-30, discount 0.15
# acct_009: grandfathered_until 2026-09-30, discount 0.10
# All others: grandfathered_until null, discount 0.0
# ---------------------------------------------------------------------------
_CONTRACTS_ALL = [
    {"account_id": "acct_001", "contract_start_date": "2024-04-01",
     "contract_end_date": "2027-03-31", "grandfathered_until": None,
     "price_floor_discount_pct": 0.0, "notes": ""},
    {"account_id": "acct_002", "contract_start_date": "2025-01-15",
     "contract_end_date": "2026-01-14", "grandfathered_until": None,
     "price_floor_discount_pct": 0.0, "notes": ""},
    {"account_id": "acct_003", "contract_start_date": "2024-07-01",
     "contract_end_date": "2027-06-30", "grandfathered_until": None,
     "price_floor_discount_pct": 0.0, "notes": ""},
    {"account_id": "acct_004", "contract_start_date": "2025-03-01",
     "contract_end_date": "2026-02-28", "grandfathered_until": None,
     "price_floor_discount_pct": 0.0, "notes": ""},
    {"account_id": "acct_005", "contract_start_date": "2025-06-01",
     "contract_end_date": "2026-05-31", "grandfathered_until": None,
     "price_floor_discount_pct": 0.0, "notes": ""},
    {"account_id": "acct_006", "contract_start_date": "2024-10-01",
     "contract_end_date": "2027-09-30", "grandfathered_until": None,
     "price_floor_discount_pct": 0.0, "notes": ""},
    {"account_id": "acct_007", "contract_start_date": "2025-07-01",
     "contract_end_date": "2027-06-30", "grandfathered_until": "2026-06-30",
     "price_floor_discount_pct": 15,
     "notes": "Eligible for temporary price floor protection"},
    {"account_id": "acct_008", "contract_start_date": "2024-02-01",
     "contract_end_date": "2027-01-31", "grandfathered_until": None,
     "price_floor_discount_pct": 0.0, "notes": ""},
    {"account_id": "acct_009", "contract_start_date": "2025-10-01",
     "contract_end_date": "2027-09-30", "grandfathered_until": "2026-09-30",
     "price_floor_discount_pct": 0.10,
     "notes": "Eligible for temporary price floor protection"},
    {"account_id": "acct_010", "contract_start_date": "2023-11-01",
     "contract_end_date": "2026-10-31", "grandfathered_until": None,
     "price_floor_discount_pct": 0.0, "notes": ""},
    {"account_id": "acct_011", "contract_start_date": "2025-04-01",
     "contract_end_date": "2027-03-31", "grandfathered_until": None,
     "price_floor_discount_pct": 0.0, "notes": ""},
    {"account_id": "acct_012", "contract_start_date": "2024-05-15",
     "contract_end_date": "2027-05-14", "grandfathered_until": None,
     "price_floor_discount_pct": 0.0, "notes": ""},
    {"account_id": "acct_013", "contract_start_date": "2025-08-01",
     "contract_end_date": "2027-07-31", "grandfathered_until": "2026-12-31",
     "price_floor_discount_pct": 0.10,
     "notes": "Eligible for temporary price floor protection"},
    {"account_id": "acct_014", "contract_start_date": "2025-09-01",
     "contract_end_date": "2027-08-31", "grandfathered_until": "2026-12-31",
     "price_floor_discount_pct": 0.15,
     "notes": "Eligible for temporary price floor protection"},
    {"account_id": "acct_015", "contract_start_date": "2025-02-01",
     "contract_end_date": "2027-01-31", "grandfathered_until": None,
     "price_floor_discount_pct": 0.0, "notes": ""},
    {"account_id": "acct_016", "contract_start_date": "2025-05-01",
     "contract_end_date": "2027-04-30", "grandfathered_until": None,
     "price_floor_discount_pct": 0.0, "notes": ""},
]
CONTRACTS = [c for c in _CONTRACTS_ALL if c["account_id"] in KEPT_ACCOUNT_IDS]

with open(os.path.join(DATA_DIR, "contract_terms.jsonl"), "w") as f:
    for c in CONTRACTS:
        f.write(json.dumps(c) + "\n")

# ---------------------------------------------------------------------------
# 4. exception_approvals.xlsx — multi-sheet formula-backed workbook
# ---------------------------------------------------------------------------
import openpyxl
from openpyxl import Workbook

wb = Workbook()
# Remove default sheet; build structured workbook
wb.remove(wb.active)

# Opaque reference sheets — names intentionally non-descriptive.
ws_policy = wb.create_sheet("Ref_R4")
for col_idx, h in enumerate(["k", "v"], start=1):
    ws_policy.cell(row=1, column=col_idx, value=h)
POLICY_ROWS = [
    ("PR_E1",  4200),
    ("PR_A8",   800),
    ("PR_EP1", 7200),
    ("PR_SF1",   75),
    # Decoys: tempting if an agent reads rate rows without evaluating Approvals formulas.
    ("PR_HALF", 2100),
    ("PR_PRE1", 3400),
    ("PR_PRE2", 3475),
    ("PR_PRE3", 3750),
]
for r_idx, row in enumerate(POLICY_ROWS, start=2):
    for c_idx, val in enumerate(row, start=1):
        ws_policy.cell(row=r_idx, column=c_idx, value=val)

ws_basis = wb.create_sheet("Ref_M3")
for col_idx, h in enumerate(["b", "rk"], start=1):
    ws_basis.cell(row=1, column=col_idx, value=h)
BASIS_ROWS = [
    ("PK01", "PR_E1"),
    ("PK02", "PR_E1"),
    ("PK03", "PR_E1"),
]
for r_idx, row in enumerate(BASIS_ROWS, start=2):
    for c_idx, val in enumerate(row, start=1):
        ws_basis.cell(row=r_idx, column=c_idx, value=val)

ws_lookup = wb.create_sheet("Ref_X7")
lookup_headers = ["approval_id", "account_id", "tgt_plan", "branch_flag", "z9"]
for col_idx, h in enumerate(lookup_headers, start=1):
    ws_lookup.cell(row=1, column=col_idx, value=h)
# tgt_plan (col 3) is a decoy for naive merges; Approvals formulas read z9 (col 5).
LOOKUP_ROWS = [
    ("EXC-001",  "acct_004", "Launch",      "PK01", "Scale"),
]
for r_idx, row in enumerate(LOOKUP_ROWS, start=2):
    for c_idx, val in enumerate(row, start=1):
        ws_lookup.cell(row=r_idx, column=c_idx, value=val)

# Decoy archive sheet — literal values that look authoritative but are stale.
ws_archive = wb.create_sheet("Approvals_Archive")
archive_headers = [
    "approval_id", "account_id", "approval_type", "approval_status",
    "effective_date", "expiry_date", "approved_plan", "approved_price",
]
for col_idx, h in enumerate(archive_headers, start=1):
    ws_archive.cell(row=1, column=col_idx, value=h)
ARCHIVE_ROWS = [
    ("EXC-001",  "acct_004", "plan_override",  "Approved", "2026-01-15", "2026-06-30", "Launch",  None,),
    ("EXC-003",  "acct_008", "price_override", "approved", "2026-02-01", "2026-12-31", None,     4200.00),
    ("EXC-003A", "acct_008", "price_override", "Superseded", "2025-11-01", "2026-12-31", None,   4200.00),
]
for r_idx, row in enumerate(ARCHIVE_ROWS, start=2):
    for c_idx, val in enumerate(row, start=1):
        ws_archive.cell(row=r_idx, column=c_idx, value=val)

ws = wb.create_sheet("Approvals")
ws["A1"] = "Pricing Exception Approval Export"
ws["A2"] = "Generated 2026-04-02"
ws["A3"] = "Internal use only"
ws["A4"] = None
# Decoy header row (human-friendly labels) — not the authoritative column header.
decoy_headers = [
    "Approval ID", "Account ID", "Exception Type", "Status",
    "Effective Date", "Expiry Date", "Approved Plan", "Approved Price",
]
for col_idx, h in enumerate(decoy_headers, start=1):
    ws.cell(row=5, column=col_idx, value=h)

# Stale quarterly export — human-friendly headers, literal stale values.
ws_stale = wb.create_sheet("Approvals_Q1")
for col_idx, h in enumerate(decoy_headers, start=1):
    ws_stale.cell(row=1, column=col_idx, value=h)
for r_idx, row in enumerate(ARCHIVE_ROWS, start=2):
    stale_row = (row[0], row[1], row[2], row[3], row[4], row[5], row[6], row[7])
    for c_idx, val in enumerate(stale_row, start=1):
        ws_stale.cell(row=r_idx, column=c_idx, value=val)

approval_headers = [
    "approval_id", "account_id", "approval_type", "approval_status",
    "effective_date", "expiry_date", "approved_plan", "approved_price",
    "b_key",
]
HEADER_ROW = 6
DATA_START = 7
for col_idx, h in enumerate(approval_headers, start=1):
    ws.cell(row=HEADER_ROW, column=col_idx, value=h)

# approval_id, account_id, approval_type, approval_status, effective_date,
# expiry_date, approved_plan, approved_price, approval_basis
APPROVALS = [
    ("EXC-001",  "acct_004", "plan_override",  "Approved",   "2026-01-15", "2026-06-30", None,    None,    "PK01"),
    ("EXC-002",  "acct_005", "plan_override",  "Expired",    "2026-01-20", "2026-03-31", "Enterprise", None, "PK03"),
    ("EXC-003A", "acct_008", "price_override", "Superseded", "2025-11-01", "2026-12-31", None,    4200.00, "PK02"),
    ("EXC-003",  "acct_008", "price_override", "approved",   "2026-02-01", "2026-12-31", None,    None,    "PK02"),
    ("EXC-005",  "acct_013", "price_override", "APPROVED",   "2025-11-01", "2027-06-30", None,    None,    "PK02"),
]

for r_idx, row in enumerate(APPROVALS, start=DATA_START):
    for c_idx, val in enumerate(row, start=1):
        ws.cell(row=r_idx, column=c_idx, value=val)

# Shorthand: rate via basis column I on the same row (two-step VLOOKUP).
def _rate(row_num: int) -> str:
    return (
        f'VLOOKUP(VLOOKUP(I{row_num},Ref_M3!A:B,2,FALSE),Ref_R4!A:B,2,FALSE)'
    )

# Formula-backed cells on Approvals (authoritative when valid).
ws.cell(row=DATA_START, column=7, value="=VLOOKUP(A7,Ref_X7!A:E,5,FALSE)")
ws.cell(row=DATA_START + 3, column=8, value=(
    f'=IF(I{DATA_START + 3}="PK02",'
    f'MAX({_rate(DATA_START + 3)}-VLOOKUP("PR_A8",Ref_R4!A:B,2,FALSE),'
    f'ROUND({_rate(DATA_START + 3)}*5/6,0)),0)'
))
ws.cell(row=DATA_START + 4, column=8, value=(
    f'=IF(I{DATA_START + 4}="PK02",'
    f'ROUNDUP({_rate(DATA_START + 4)}*82738/100000,-2),0)'
))

wb.save(os.path.join(DATA_DIR, "exception_approvals.xlsx"))

# ---------------------------------------------------------------------------
# 5. migration_rules.pdf  (authoritative rule document via reportlab)
# ---------------------------------------------------------------------------
from reportlab.lib.pagesizes import LETTER
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib import colors

pdf_path = os.path.join(DATA_DIR, "migration_rules.pdf")
doc = SimpleDocTemplate(pdf_path, pagesize=LETTER,
                        leftMargin=1*inch, rightMargin=1*inch,
                        topMargin=1*inch, bottomMargin=1*inch)
styles = getSampleStyleSheet()
h1 = styles["Heading1"]
h2 = styles["Heading2"]
body = styles["BodyText"]
body.spaceAfter = 6
code_style = ParagraphStyle("Code", parent=styles["BodyText"],
                             fontName="Courier", fontSize=9, spaceAfter=4)

def p(text, style=None):
    return Paragraph(text, style or body)

def section(title):
    return [Spacer(1, 12), Paragraph(title, h2)]

def kv_table(rows):
    data = [["Parameter", "Value"]] + rows
    t = Table(data, colWidths=[2.8*inch, 3.4*inch])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.lightblue),
        ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID",       (0, 0), (-1, -1), 0.5, colors.grey),
        ("FONTSIZE",   (0, 0), (-1, -1), 9),
        ("VALIGN",     (0, 0), (-1, -1), "TOP"),
    ]))
    return t

story = []

# Title
story.append(Paragraph("Pricing Migration Rule Document", styles["Title"]))
story.append(Paragraph("SaaS 2026 Plan Migration — Authoritative Reference", styles["Heading3"]))
story.append(Spacer(1, 12))
story.append(p("This document defines all rules governing the migration of SaaS customer accounts "
               "from legacy plans to new 2026 pricing packages. All audit tools must use this "
               "document as the sole source of truth for plan mapping, price floors, exception "
               "handling, grandfathering, rule precedence, and output schemas."))

# Section 1
story += section("1. Migration Effective Date")
story.append(kv_table([["Migration effective date", "2026-04-01"]]))

# Section 2
story += section("2. Usage Tier Calculation")
story.append(p("Usage tier is determined from trailing 90-day API call volume ending before "
               "the migration effective date. Events on or after the migration effective date "
               "are excluded."))

# Section 3
story += section("3. Usage Tier Thresholds")
story.append(kv_table([
    ["Low",    "Trailing 90-day API calls < 100,000"],
    ["Medium", "Trailing 90-day API calls >= 100,000 and < 500,000"],
    ["High",   "Trailing 90-day API calls >= 500,000"],
]))

# Section 4
story += section("4. Default Plan Mapping")
story.append(p("The default new plan is determined by the account's legacy plan and usage tier:"))
mapping_data = [
    ["Legacy Plan",         "Usage Tier", "New Plan"],
    ["Legacy Starter",      "Low",        "Launch"],
    ["Legacy Starter",      "Medium",     "Scale"],
    ["Legacy Starter",      "High",       "Scale Plus"],
    ["Legacy Growth",       "Low",        "Scale"],
    ["Legacy Growth",       "Medium",     "Scale Plus"],
    ["Legacy Growth",       "High",       "Enterprise"],
    ["Legacy Pro",          "Low",        "Scale Plus"],
    ["Legacy Pro",          "Medium",     "Enterprise"],
    ["Legacy Pro",          "High",       "Enterprise Plus"],
    ["Legacy Enterprise",   "Any",        "Enterprise Plus"],
]
mt = Table(mapping_data, colWidths=[2*inch, 1.5*inch, 2*inch])
mt.setStyle(TableStyle([
    ("BACKGROUND", (0, 0), (-1, 0), colors.lightblue),
    ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
    ("GRID",       (0, 0), (-1, -1), 0.5, colors.grey),
    ("FONTSIZE",   (0, 0), (-1, -1), 9),
]))
story.append(mt)

# Section 5
story += section("5. Base Monthly Prices")
story.append(kv_table([
    ["Launch",          "$500 per month"],
    ["Scale",           "$1,200 per month"],
    ["Scale Plus",      "$2,200 per month"],
    ["Enterprise",      "$4,200 per month"],
    ["Enterprise Plus", "$7,200 per month"],
]))

# Section 6
story += section("6. Seat Price Floors")
story.append(kv_table([
    ["Launch",          "$50 per seat per month"],
    ["Scale",           "$60 per seat per month"],
    ["Scale Plus",      "$75 per seat per month"],
    ["Enterprise",      "$90 per seat per month"],
    ["Enterprise Plus", "$110 per seat per month"],
]))

# Section 7
story += section("7. Expected Monthly Price Calculation")
story.append(p("For each account, the expected monthly price is the greater of two amounts: "
               "the plan's published base monthly rate (Section 5) and the account's total "
               "seat-count obligation — current seats multiplied by the per-seat floor for "
               "the applicable plan (Section 6)."))
story.append(p("Round monetary values to 2 decimal places."))

# Section 8
story += section("8. Grandfathering Rules")
story.append(p("Some accounts have a grandfathered_until date and a price_floor_discount_pct "
               "in their contract terms."))
story.append(p("Grandfathering applies to the per-seat price floor only, not the base "
               "monthly rate or plan mapping."))
story.append(p("When grandfathered_until is on or after migration_effective_date, apply "
               "price_floor_discount_pct to the per-seat floor before computing the "
               "seat-count obligation. Normalize contract field encodings before use."))
story.append(p("If grandfathered_until is null or before migration_effective_date, "
               "grandfathering does not apply."))

# Section 9
story += section("9. Exception Approval Rules")
story.append(p("Exception approvals are stored in exception_approvals.xlsx."))
story.append(p("An approval is valid when approval_status is approved (after "
               "normalizing case and surrounding whitespace), effective_date is on "
               "or before migration_effective_date, and expiry_date is on or after "
               "migration_effective_date."))
story.append(p("A valid plan_override replaces the mapped plan. A valid price_override "
               "replaces the computed monthly price."))

# Section 10
story += section("10. Rule Precedence")
story.append(p("Determine usage tier, then plan mapping, then apply valid plan overrides, "
               "then compute price from seat floors and grandfathering, then apply valid "
               "price overrides."))

# Section 11
story += section("11. Violation Types")
story.append(p("Each account is assigned exactly one violation_type:"))
story.append(kv_table([
    ["no_violation",             "Expected plan and price match assignment."],
    ["usage_tier_miscalculated", "Assigned plan matches wrong usage tier."],
    ["wrong_plan",               "Assigned plan does not match expected plan, not due to usage tier."],
    ["wrong_price_floor",        "Plan is correct, but assigned monthly price does not match expected monthly price, and the mismatch is not better explained by an expired exception or missing grandfathering."],
    ["missing_grandfathering",   "Plan is correct, but assigned price does not reflect the grandfathering-adjusted seat floor used to compute the expected monthly price."],
    ["expired_exception",        "Assignment relied on an approval whose expiry_date is before migration_effective_date."],
]))

# Section 12
story += section("12. Required Output Schemas")
story.append(p("<b>migration_audit.csv</b> — one row per account:"))
story.append(Paragraph(
    "account_id | legacy_plan | usage_tier | expected_new_plan | assigned_new_plan | "
    "expected_monthly_price | assigned_monthly_price | migration_status | violation_type | "
    "revenue_impact_usd",
    code_style))
story.append(p("migration_status allowed values: pass, fail"))
story.append(p("violation_type allowed values: no_violation, usage_tier_miscalculated, "
               "wrong_plan, wrong_price_floor, missing_grandfathering, expired_exception"))
story.append(p("revenue_impact_usd = assigned_monthly_price - expected_monthly_price "
               "(positive = overcharged, negative = undercharged, zero = pass)"))

story.append(Spacer(1, 6))
story.append(p("<b>rule_violations.json</b> — summary object with keys:"))
story.append(Paragraph(
    "total_accounts, failed_accounts, passed_accounts, violation_counts "
    "(dict with keys: wrong_plan, wrong_price_floor, expired_exception, "
    "missing_grandfathering, usage_tier_miscalculated), "
    "highest_revenue_impact_account, net_revenue_impact_usd, policy_sources",
    code_style))
story.append(p("policy_sources must list: migration_rules.pdf, legacy_accounts.csv, "
               "product_usage_events.csv, contract_terms.jsonl, exception_approvals.xlsx, "
               "new_plan_assignments.csv"))

story.append(Spacer(1, 6))
story.append(p("<b>revenue_impact_summary.csv</b> — rows with metric and value columns:"))
story.append(Paragraph(
    "total_overcharge_usd | total_undercharge_usd | net_revenue_impact_usd | "
    "absolute_revenue_impact_usd | failed_account_count",
    code_style))
story.append(p("total_overcharge_usd: sum of positive revenue_impact_usd values."))
story.append(p("total_undercharge_usd: absolute sum of negative revenue_impact_usd values (positive number)."))
story.append(p("All metric values must be numeric."))

story.append(Spacer(1, 6))
story.append(p("<b>exception_review.csv</b> — one row per approval record:"))
story.append(Paragraph(
    "account_id | approval_id | exception_type | expiry_date | "
    "is_valid_on_migration_date | applied_to_expected_result | review_status",
    code_style))
story.append(p("review_status allowed values: valid_applied, expired_not_applied, "
               "superseded_not_applied, not_applicable"))

doc.build(story)

# ---------------------------------------------------------------------------
# 6. new_plan_assignments.csv
# Contains the actual (partially wrong) migration output from the company system.
# ---------------------------------------------------------------------------
_NEW_ASSIGNMENTS_ALL = [
    # account_id, assigned_new_plan, assigned_monthly_price, loaded_by
    ("acct_001", "Scale",           1200.00, "migration_batch"),
    ("acct_002", "Enterprise",      4200.00, "migration_batch"),
    ("acct_003", "Scale",           1200.00, "migration_batch"),
    ("acct_004", "Scale",           1200.00, "migration_batch"),
    ("acct_005", "Enterprise",      4200.00, "migration_batch"),
    ("acct_006", "Scale",           1200.00, "migration_batch"),
    ("acct_007", "Scale Plus",      2625.00, "migration_batch"),
    ("acct_008", "Enterprise",      3500.00, "migration_batch"),
    ("acct_009", "Launch",          500.00, "migration_batch"),
    ("acct_010", "Enterprise Plus", 7200.00, "migration_batch"),
    ("acct_011", "Scale",           1680.00, "migration_batch"),
    ("acct_012", "Enterprise Plus", 3800.00, "migration_batch"),
    ("acct_013", "Enterprise",      3150.00, "migration_batch"),
    ("acct_014", "Scale Plus",      2550.00, "migration_batch"),
    ("acct_015", "Enterprise",      3900.00, "migration_batch"),
    ("acct_016", "Enterprise",      3780.00, "migration_batch"),
]
NEW_ASSIGNMENTS = [r for r in _NEW_ASSIGNMENTS_ALL if _keep_account(r)]

with open(os.path.join(DATA_DIR, "new_plan_assignments.csv"), "w", newline="") as f:
    w = csv.writer(f)
    w.writerow(["account_id", "assigned_new_plan", "assigned_monthly_price", "loaded_by"])
    w.writerows(NEW_ASSIGNMENTS)

# ---------------------------------------------------------------------------
# 7. Buggy project files under /root/project
# ---------------------------------------------------------------------------

# ---- config.yaml -----------------------------------------------------------
config_yaml = """\
data_dir: /root/data
out_dir: /root/out
migration_effective_date: "2026-04-01"
"""

with open(os.path.join(PROJECT_DIR, "config.yaml"), "w", encoding="utf-8") as f:
    f.write(config_yaml)

# ---- rules.py  (buggy) -----------------------------------------------------
# Defects (unlabeled): uses latest calendar month for usage (not trailing 90-day);
# no expiry check on exceptions; grandfathering incorrectly applied to plan family.
rules_py = '''\
"""rules.py - Migration plan assignment and exception processing."""

import pandas as pd

PLAN_MAPPING = {
    ("Legacy Starter",    "Low"):    "Launch",
    ("Legacy Starter",    "Medium"): "Scale",
    ("Legacy Starter",    "High"):   "Scale Plus",
    ("Legacy Growth",     "Low"):    "Scale",
    ("Legacy Growth",     "Medium"): "Scale Plus",
    ("Legacy Growth",     "High"):   "Enterprise",
    ("Legacy Pro",        "Low"):    "Scale Plus",
    ("Legacy Pro",        "Medium"): "Enterprise",
    ("Legacy Pro",        "High"):   "Enterprise Plus",
    ("Legacy Enterprise", "Low"):    "Enterprise Plus",
    ("Legacy Enterprise", "Medium"): "Enterprise Plus",
    ("Legacy Enterprise", "High"):   "Enterprise Plus",
}

PLAN_HIERARCHY = ["Launch", "Scale", "Scale Plus", "Enterprise", "Enterprise Plus"]


def calculate_usage_tier(events_df, account_id, migration_effective_date):
    """Compute usage tier for an account."""
    acct_events = events_df[events_df["account_id"] == account_id]
    if acct_events.empty:
        return "Low"
    latest_month = acct_events["event_date"].dt.to_period("M").max()
    monthly = acct_events[acct_events["event_date"].dt.to_period("M") == latest_month]
    total_calls = int(monthly["api_calls"].sum())
    if total_calls >= 500_000:
        return "High"
    elif total_calls >= 100_000:
        return "Medium"
    return "Low"


def _get_default_plan(legacy_plan, usage_tier):
    return PLAN_MAPPING.get((legacy_plan, usage_tier), "Scale")


def apply_exception(account_id, default_plan, approvals_df, migration_effective_date):
    """Apply plan exception approvals for the account."""
    if approvals_df is None or approvals_df.empty:
        return default_plan
    acct_approvals = approvals_df[
        (approvals_df["account_id"] == account_id)
        & (approvals_df["exception_type"] == "plan_override")
    ]
    if acct_approvals.empty:
        return default_plan
    row = acct_approvals.iloc[0]
    approved_plan = row.get("approved_plan")
    if pd.notna(approved_plan) and str(approved_plan).strip():
        return str(approved_plan).strip()
    return default_plan


def apply_grandfathering(account_id, plan, contracts_df, migration_effective_date):
    """Apply contract grandfathering terms."""
    if contracts_df is None or contracts_df.empty:
        return plan
    contract = contracts_df[contracts_df["account_id"] == account_id]
    if contract.empty:
        return plan
    row = contract.iloc[0]
    gf_until = row.get("grandfathered_until")
    if not gf_until or pd.isna(gf_until):
        return plan
    try:
        gf_dt = pd.Timestamp(gf_until)
        migration_dt = pd.Timestamp(migration_effective_date)
        if gf_dt >= migration_dt and plan in PLAN_HIERARCHY:
            idx = PLAN_HIERARCHY.index(plan)
            if idx > 0:
                return PLAN_HIERARCHY[idx - 1]
    except Exception:
        pass
    return plan


def determine_expected_plan(account_id, legacy_plan, events_df, approvals_df,
                            contracts_df, migration_effective_date):
    """Determine the expected new plan for an account."""
    usage_tier = calculate_usage_tier(events_df, account_id, migration_effective_date)
    default_plan = _get_default_plan(legacy_plan, usage_tier)
    plan_with_exception = apply_exception(
        account_id, default_plan, approvals_df, migration_effective_date
    )
    final_plan = apply_grandfathering(
        account_id, plan_with_exception, contracts_df, migration_effective_date
    )
    return usage_tier, final_plan
'''

with open(os.path.join(PROJECT_DIR, "rules.py"), "w", encoding="utf-8") as f:
    f.write(rules_py)

# ---- pricing.py  (buggy) ---------------------------------------------------
# Defects (unlabeled): grandfathering discount applied to base price allocation
# instead of per-seat floor rate; Scale plan ignores seat-count floor entirely.
pricing_py = '''\
"""pricing.py - Price floor and monthly price calculations."""

import pandas as pd

BASE_PRICES = {
    "Launch":          500.0,
    "Scale":           1200.0,
    "Scale Plus":      2200.0,
    "Enterprise":      4200.0,
    "Enterprise Plus": 7200.0,
}

SEAT_FLOORS = {
    "Launch":          50.0,
    "Scale":           60.0,
    "Scale Plus":      75.0,
    "Enterprise":      90.0,
    "Enterprise Plus": 110.0,
}


def calculate_price_floor(plan, seats, contract_row, migration_effective_date):
    """Calculate the applicable seat price floor for the account\'s plan."""
    seat_floor = SEAT_FLOORS.get(plan, 0.0)
    discount = 0.0
    if contract_row is not None:
        gf_until = contract_row.get("grandfathered_until")
        if gf_until and not (isinstance(gf_until, float) and pd.isna(gf_until)):
            try:
                gf_dt = pd.Timestamp(gf_until)
                migration_dt = pd.Timestamp(migration_effective_date)
                if gf_dt >= migration_dt:
                    discount = float(contract_row.get("price_floor_discount_pct", 0.0) or 0.0)
            except Exception:
                pass
    if discount > 0:
        return (BASE_PRICES.get(plan, 0.0) * (1.0 - discount)) / max(seats, 1)
    if plan == "Scale":
        return 0.0
    return seat_floor


def calculate_expected_price(plan, seats, contract_row, migration_effective_date):
    """Calculate the expected monthly price from plan, seat count, and contract terms."""
    floor_per_seat = calculate_price_floor(plan, seats, contract_row, migration_effective_date)
    base = BASE_PRICES.get(plan, 0.0)
    return round(max(base, seats * floor_per_seat), 2)


def apply_price_exception(account_id, computed_price, approvals_df, migration_effective_date):
    """Apply price exception approvals for the account."""
    if approvals_df is None or approvals_df.empty:
        return computed_price
    acct_approvals = approvals_df[
        (approvals_df["account_id"] == account_id)
        & (approvals_df["exception_type"] == "price_override")
    ]
    if acct_approvals.empty:
        return computed_price
    row = acct_approvals.iloc[0]
    approved_price = row.get("approved_price")
    if pd.notna(approved_price):
        return round(float(approved_price), 2)
    return computed_price
'''

with open(os.path.join(PROJECT_DIR, "pricing.py"), "w", encoding="utf-8") as f:
    f.write(pricing_py)

# ---- audit_migration.py  (buggy) -------------------------------------------
# Defects (unlabeled): Excel header read with row 0 instead of actual header row;
# revenue impact sign reversed; rule_violations.json missing required summary keys.
audit_migration_py = '''\
"""audit_migration.py - Main migration audit orchestration script."""

import argparse
import json
import os
import sys

import pandas as pd
import yaml


def load_config(config_path):
    with open(config_path) as f:
        return yaml.safe_load(f)


def load_exception_approvals(data_dir):
    """Load exception approvals from the approvals workbook."""
    path = os.path.join(data_dir, "exception_approvals.xlsx")
    try:
        df = pd.read_excel(path, sheet_name="Approvals", header=0)
        return df[["approval_id", "account_id", "approval_type",
                   "approved_plan", "approved_price", "effective_date",
                   "expiry_date", "approval_status"]]
    except (KeyError, ValueError):
        return pd.DataFrame(columns=["approval_id", "account_id", "approval_type",
                                     "approved_plan", "approved_price",
                                     "effective_date", "expiry_date", "approval_status"])


def run_audit(config_path):
    cfg = load_config(config_path)
    data_dir = cfg["data_dir"]
    out_dir = cfg["out_dir"]
    migration_date = cfg["migration_effective_date"]

    os.makedirs(out_dir, exist_ok=True)

    # Load inputs
    accounts_df = pd.read_csv(os.path.join(data_dir, "legacy_accounts.csv"))
    events_df = pd.read_csv(os.path.join(data_dir, "product_usage_events.csv"),
                            parse_dates=["event_date"])
    contracts = {}
    with open(os.path.join(data_dir, "contract_terms.jsonl")) as f:
        for line in f:
            rec = json.loads(line.strip())
            contracts[rec["account_id"]] = rec
    approvals_df = load_exception_approvals(data_dir)
    assignments_df = pd.read_csv(os.path.join(data_dir, "new_plan_assignments.csv"))

    sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
    from rules import determine_expected_plan
    from pricing import calculate_expected_price, apply_price_exception

    contracts_df = pd.DataFrame(contracts.values())
    audit_rows = []
    for _, acct in accounts_df.iterrows():
        account_id = acct["account_id"]
        legacy_plan = acct["legacy_plan"]
        seats = int(acct["current_seats"])
        contract_row = contracts.get(account_id)

        usage_tier, expected_plan = determine_expected_plan(
            account_id, legacy_plan, events_df, approvals_df,
            contracts_df=contracts_df,
            migration_effective_date=migration_date
        )

        computed_price = calculate_expected_price(
            expected_plan, seats, contract_row, migration_date
        )
        expected_price = apply_price_exception(
            account_id, computed_price, approvals_df, migration_date
        )

        assignment = assignments_df[assignments_df["account_id"] == account_id]
        assigned_plan = assignment["assigned_new_plan"].values[0]
        assigned_price = float(assignment["assigned_monthly_price"].values[0])

        # Determine violation type
        violation_type = _determine_violation(
            account_id, usage_tier, expected_plan, assigned_plan,
            expected_price, assigned_price, approvals_df, contracts,
            migration_date
        )

        migration_status = "pass" if violation_type == "no_violation" else "fail"

        revenue_impact = round(expected_price - assigned_price, 2)

        audit_rows.append({
            "account_id": account_id,
            "legacy_plan": legacy_plan,
            "usage_tier": usage_tier,
            "expected_new_plan": expected_plan,
            "assigned_new_plan": assigned_plan,
            "expected_monthly_price": expected_price,
            "assigned_monthly_price": assigned_price,
            "migration_status": migration_status,
            "violation_type": violation_type,
            "revenue_impact_usd": revenue_impact,
        })

    audit_df = pd.DataFrame(audit_rows)
    audit_df.to_csv(os.path.join(out_dir, "migration_audit.csv"), index=False)

    failed = audit_df[audit_df["migration_status"] == "fail"]
    passed = audit_df[audit_df["migration_status"] == "pass"]
    violation_counts = failed["violation_type"].value_counts().to_dict()

    violations_summary = {
        "total_accounts": len(audit_df),
        "failed_accounts": len(failed),
        "passed_accounts": len(passed),
        "violation_counts": {
            "wrong_plan": violation_counts.get("wrong_plan", 0),
            "wrong_price_floor": violation_counts.get("wrong_price_floor", 0),
        },
    }

    with open(os.path.join(out_dir, "rule_violations.json"), "w") as f:
        json.dump(violations_summary, f, indent=2)

    # revenue_impact_summary.csv
    overcharges = audit_df[audit_df["revenue_impact_usd"] > 0]["revenue_impact_usd"]
    undercharges = audit_df[audit_df["revenue_impact_usd"] < 0]["revenue_impact_usd"]
    summary_rows = [
        {"metric": "total_overcharge_usd",
         "value": round(overcharges.sum(), 2)},
        {"metric": "total_undercharge_usd",
         "value": round(abs(undercharges.sum()), 2)},
        {"metric": "net_revenue_impact_usd",
         "value": round(audit_df["revenue_impact_usd"].sum(), 2)},
        {"metric": "absolute_revenue_impact_usd",
         "value": round(audit_df["revenue_impact_usd"].abs().sum(), 2)},
        {"metric": "failed_account_count",
         "value": len(failed)},
    ]
    pd.DataFrame(summary_rows).to_csv(
        os.path.join(out_dir, "revenue_impact_summary.csv"), index=False
    )

    # exception_review.csv
    review_rows = _build_exception_review(
        approvals_df, audit_df, migration_date
    )
    pd.DataFrame(review_rows).to_csv(
        os.path.join(out_dir, "exception_review.csv"), index=False
    )


def _determine_violation(account_id, usage_tier, expected_plan, assigned_plan,
                         expected_price, assigned_price, approvals_df,
                         contracts, migration_date):
    """Determine the primary violation type for an account."""
    if assigned_plan != expected_plan:
        return "wrong_plan"
    if abs(assigned_price - expected_price) > 0.01:
        return "wrong_price_floor"
    return "no_violation"


def _build_exception_review(approvals_df, audit_df, migration_date):
    rows = []
    if approvals_df is None or approvals_df.empty:
        return rows
    for _, row in approvals_df.iterrows():
        account_id = row["account_id"]
        approval_id = row["approval_id"]
        exc_type = row.get("approval_type", row.get("exception_type", ""))
        expiry = row.get("expiry_date")
        try:
            is_valid = pd.Timestamp(expiry) >= pd.Timestamp(migration_date)
        except Exception:
            is_valid = True
        rows.append({
            "account_id": account_id,
            "approval_id": approval_id,
            "exception_type": exc_type,
            "expiry_date": str(expiry)[:10] if pd.notna(expiry) else "",
            "is_valid_on_migration_date": is_valid,
            "applied_to_expected_result": is_valid,
            "review_status": "valid_applied" if is_valid else "not_applicable",
        })
    return rows


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--config", required=True)
    args = parser.parse_args()
    run_audit(args.config)
'''

with open(os.path.join(PROJECT_DIR, "audit_migration.py"), "w", encoding="utf-8") as f:
    f.write(audit_migration_py)

print("build_inputs.py complete — all files written.")
print(f"  Data files: {os.listdir(DATA_DIR)}")
print(f"  Project files: {os.listdir(PROJECT_DIR)}")
