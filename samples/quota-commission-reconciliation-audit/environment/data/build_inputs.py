"""
Generate deterministic input files for quota-commission-reconciliation-audit.
All expected verifier values are derivable from constants below.
"""

import json
import math
import os

import pandas as pd
from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule
from openpyxl.styles import Font, PatternFill
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas

OUTPUT_DIR = "/root/data"
os.makedirs(OUTPUT_DIR, exist_ok=True)

# base_quota in USD; carry_rate as decimal; formula ROUND or MAX (REP-012, REP-015)
# Comments show expected tier after full audit (bookings designed per tier targets).
REPS = [
    # Below Threshold ~70%
    {"rep_id": "REP-001", "rep_name": "Jordan Ellis", "region": "NA-East",
     "base_quota": 145000, "carry_rate": 0.10, "formula": "ROUND"},  # Accel 1, 105%
    {"rep_id": "REP-002", "rep_name": "Morgan Reed", "region": "NA-West",
     "base_quota": 120000, "carry_rate": 0.08, "formula": "ROUND"},  # Below, 70%
    {"rep_id": "REP-003", "rep_name": "Priya Sharma", "region": "EMEA",
     "base_quota": 180000, "carry_rate": 0.12, "formula": "ROUND"},  # Accel 3, 165%
    {"rep_id": "REP-004", "rep_name": "Kenji Tanaka", "region": "APAC",
     "base_quota": 100000, "carry_rate": 0.05, "formula": "ROUND"},  # Base, 90%
    {"rep_id": "REP-005", "rep_name": "Sofia Alvarez", "region": "NA-East",
     "base_quota": 200000, "carry_rate": 0.15, "formula": "ROUND"},  # Accel 2, 135%
    {"rep_id": "REP-006", "rep_name": "Liam O'Brien", "region": "LATAM",
     "base_quota": 40000, "carry_rate": 0.10, "formula": "ROUND"},  # Below, 76.7%
    {"rep_id": "REP-007", "rep_name": "Aisha Khan", "region": "EMEA",
     "base_quota": 160000, "carry_rate": 0.10, "formula": "ROUND"},  # Accel 1, 110%
    {"rep_id": "REP-008", "rep_name": "Tyler Brooks", "region": "NA-West",
     "base_quota": 95000, "carry_rate": 0.00, "formula": "ROUND"},  # Base, 88%; zero carry
    {"rep_id": "REP-009", "rep_name": "Nina Patel", "region": "APAC",
     "base_quota": 55000, "carry_rate": 0.07, "formula": "ROUND"},  # Base, 85%
    {"rep_id": "REP-010", "rep_name": "Carlos Mendez", "region": "LATAM",
     "base_quota": 75000, "carry_rate": 0.09, "formula": "ROUND"},  # Below, 70.1%
    {"rep_id": "REP-011", "rep_name": "Emily Chen", "region": "NA-East",
     "base_quota": 110000, "carry_rate": 0.11, "formula": "ROUND"},  # Accel 1, 105%
    {"rep_id": "REP-012", "rep_name": "Marcus Webb", "region": "NA-West",
     "base_quota": 65000, "carry_rate": 0.06, "formula": "MAX"},  # Accel 2 ~148%; MAX cascade trap
    {"rep_id": "REP-013", "rep_name": "Olivia Grant", "region": "EMEA",
     "base_quota": 220000, "carry_rate": 0.15, "formula": "ROUND"},  # Accel 3, 160%
    {"rep_id": "REP-014", "rep_name": "David Kim", "region": "APAC",
     "base_quota": 85000, "carry_rate": 0.08, "formula": "ROUND"},  # Base, 92%
    {"rep_id": "REP-015", "rep_name": "Rachel Torres", "region": "LATAM",
     "base_quota": 50000, "carry_rate": 0.05, "formula": "MAX"},  # Below ~72%; MAX trap
    {"rep_id": "REP-016", "rep_name": "James Liu", "region": "NA-East",
     "base_quota": 135000, "carry_rate": 0.12, "formula": "ROUND"},  # Accel 1, 115%
    {"rep_id": "REP-017", "rep_name": "Hannah Ross", "region": "EMEA",
     "base_quota": 70000, "carry_rate": 0.10, "formula": "ROUND"},  # Accel 2, 140%
    {"rep_id": "REP-018", "rep_name": "Alex Novak", "region": "NA-West",
     "base_quota": 48000, "carry_rate": 0.07, "formula": "ROUND"},  # Base, 88%
    {"rep_id": "REP-019", "rep_name": "Sara Mohammed", "region": "APAC",
     "base_quota": 175000, "carry_rate": 0.13, "formula": "ROUND"},  # Accel 3, 155%
    {"rep_id": "REP-020", "rep_name": "Ben Carter", "region": "LATAM",
     "base_quota": 60000, "carry_rate": 0.00, "formula": "ROUND"},  # Below, 60%; zero carry
]

# Target bookings_arr_usd per rep (deals below sum to these totals).
BOOKINGS_TARGET = {
    "REP-001": 168000.0,
    "REP-002": 91000.0,
    "REP-003": 333300.0,
    "REP-004": 94500.0,
    "REP-005": 310500.0,
    "REP-006": 33750.0,
    "REP-007": 193600.0,
    "REP-008": 83600.0,
    "REP-009": 50200.0,
    "REP-010": 57500.0,
    "REP-011": 128100.0,
    "REP-012": 103500.0,
    "REP-013": 404800.0,
    "REP-014": 84640.0,
    "REP-015": 39750.0,
    "REP-016": 173650.0,
    "REP-017": 107800.0,
    "REP-018": 44880.0,
    "REP-019": 306900.0,
    "REP-020": 36000.0,
}

# deal_id -> formatted currency string for arr_usd (Trap C: mixed JSONL types)
STRING_ARR_DEALS = {
    "D-001-02": "$35,000.00",
    "D-003-03": "$55,000.00",
    "D-011-02": "$42,000.00",
    "D-015-01": "$22,000.00",
}

DEALS = [
    {"deal_id": "D-001-01", "rep_id": "REP-001", "close_date": "2026-01-15", "arr_usd": 42000.0},
    {"deal_id": "D-001-02", "rep_id": "REP-001", "close_date": "2026-01-28", "arr_usd": 35000.0},
    {"deal_id": "D-001-03", "rep_id": "REP-001", "close_date": "2026-02-10", "arr_usd": 38000.0},
    {"deal_id": "D-001-04", "rep_id": "REP-001", "close_date": "2026-02-22", "arr_usd": 28000.0},
    {"deal_id": "D-001-05", "rep_id": "REP-001", "close_date": "2026-03-08", "arr_usd": 25000.0},
    {"deal_id": "D-002-01", "rep_id": "REP-002", "close_date": "2026-01-12", "arr_usd": 22000.0},
    {"deal_id": "D-002-02", "rep_id": "REP-002", "close_date": "2026-02-05", "arr_usd": 24000.0},
    {"deal_id": "D-002-03", "rep_id": "REP-002", "close_date": "2026-02-18", "arr_usd": 21000.0},
    {"deal_id": "D-002-04", "rep_id": "REP-002", "close_date": "2026-03-14", "arr_usd": 24000.0},
    {"deal_id": "D-003-01", "rep_id": "REP-003", "close_date": "2026-01-08", "arr_usd": 58000.0},
    {"deal_id": "D-003-02", "rep_id": "REP-003", "close_date": "2026-01-25", "arr_usd": 62000.0},
    {"deal_id": "D-003-03", "rep_id": "REP-003", "close_date": "2026-02-12", "arr_usd": 55000.0},
    {"deal_id": "D-003-04", "rep_id": "REP-003", "close_date": "2026-02-27", "arr_usd": 48000.0},
    {"deal_id": "D-003-05", "rep_id": "REP-003", "close_date": "2026-03-05", "arr_usd": 55300.0},
    {"deal_id": "D-003-06", "rep_id": "REP-003", "close_date": "2026-03-20", "arr_usd": 55000.0},
    {"deal_id": "D-004-01", "rep_id": "REP-004", "close_date": "2026-01-20", "arr_usd": 25000.0},
    {"deal_id": "D-004-02", "rep_id": "REP-004", "close_date": "2026-02-08", "arr_usd": 22000.0},
    {"deal_id": "D-004-03", "rep_id": "REP-004", "close_date": "2026-03-01", "arr_usd": 24500.0},
    {"deal_id": "D-004-04", "rep_id": "REP-004", "close_date": "2026-03-18", "arr_usd": 23000.0},
    {"deal_id": "D-005-01", "rep_id": "REP-005", "close_date": "2026-01-10", "arr_usd": 65000.0},
    {"deal_id": "D-005-02", "rep_id": "REP-005", "close_date": "2026-01-30", "arr_usd": 72000.0},
    {"deal_id": "D-005-03", "rep_id": "REP-005", "close_date": "2026-02-15", "arr_usd": 58500.0},
    {"deal_id": "D-005-04", "rep_id": "REP-005", "close_date": "2026-03-03", "arr_usd": 58000.0},
    {"deal_id": "D-005-05", "rep_id": "REP-005", "close_date": "2026-03-25", "arr_usd": 57000.0},
    {"deal_id": "D-006-01", "rep_id": "REP-006", "close_date": "2026-01-17", "arr_usd": 8500.0},
    {"deal_id": "D-006-02", "rep_id": "REP-006", "close_date": "2026-02-02", "arr_usd": 9000.0},
    {"deal_id": "D-006-03", "rep_id": "REP-006", "close_date": "2026-02-20", "arr_usd": 7750.0},
    {"deal_id": "D-006-04", "rep_id": "REP-006", "close_date": "2026-03-12", "arr_usd": 8500.0},
    {"deal_id": "D-007-01", "rep_id": "REP-007", "close_date": "2026-01-14", "arr_usd": 52000.0},
    {"deal_id": "D-007-02", "rep_id": "REP-007", "close_date": "2026-02-09", "arr_usd": 48600.0},
    {"deal_id": "D-007-03", "rep_id": "REP-007", "close_date": "2026-03-06", "arr_usd": 46800.0},
    {"deal_id": "D-007-04", "rep_id": "REP-007", "close_date": "2026-03-22", "arr_usd": 46200.0},
    {"deal_id": "D-008-01", "rep_id": "REP-008", "close_date": "2026-01-22", "arr_usd": 28600.0},
    {"deal_id": "D-008-02", "rep_id": "REP-008", "close_date": "2026-02-14", "arr_usd": 27500.0},
    {"deal_id": "D-008-03", "rep_id": "REP-008", "close_date": "2026-03-10", "arr_usd": 27500.0},
    {"deal_id": "D-009-01", "rep_id": "REP-009", "close_date": "2026-01-11", "arr_usd": 12800.0},
    {"deal_id": "D-009-02", "rep_id": "REP-009", "close_date": "2026-02-03", "arr_usd": 12400.0},
    {"deal_id": "D-009-03", "rep_id": "REP-009", "close_date": "2026-02-19", "arr_usd": 12600.0},
    {"deal_id": "D-009-04", "rep_id": "REP-009", "close_date": "2026-03-07", "arr_usd": 12400.0},
    {"deal_id": "D-010-01", "rep_id": "REP-010", "close_date": "2026-01-16", "arr_usd": 14500.0},
    {"deal_id": "D-010-02", "rep_id": "REP-010", "close_date": "2026-02-06", "arr_usd": 14200.0},
    {"deal_id": "D-010-03", "rep_id": "REP-010", "close_date": "2026-02-21", "arr_usd": 14300.0},
    {"deal_id": "D-010-04", "rep_id": "REP-010", "close_date": "2026-03-13", "arr_usd": 14500.0},
    {"deal_id": "D-011-01", "rep_id": "REP-011", "close_date": "2026-01-09", "arr_usd": 32100.0},
    {"deal_id": "D-011-02", "rep_id": "REP-011", "close_date": "2026-01-27", "arr_usd": 42000.0},
    {"deal_id": "D-011-03", "rep_id": "REP-011", "close_date": "2026-02-11", "arr_usd": 28000.0},
    {"deal_id": "D-011-04", "rep_id": "REP-011", "close_date": "2026-03-04", "arr_usd": 26000.0},
    {"deal_id": "D-012-01", "rep_id": "REP-012", "close_date": "2026-01-18", "arr_usd": 25875.0},
    {"deal_id": "D-012-02", "rep_id": "REP-012", "close_date": "2026-02-04", "arr_usd": 25875.0},
    {"deal_id": "D-012-03", "rep_id": "REP-012", "close_date": "2026-02-17", "arr_usd": 25875.0},
    {"deal_id": "D-012-04", "rep_id": "REP-012", "close_date": "2026-03-09", "arr_usd": 25875.0},
    {"deal_id": "D-013-01", "rep_id": "REP-013", "close_date": "2026-01-07", "arr_usd": 81000.0},
    {"deal_id": "D-013-02", "rep_id": "REP-013", "close_date": "2026-01-24", "arr_usd": 80800.0},
    {"deal_id": "D-013-03", "rep_id": "REP-013", "close_date": "2026-02-10", "arr_usd": 81000.0},
    {"deal_id": "D-013-04", "rep_id": "REP-013", "close_date": "2026-02-26", "arr_usd": 80800.0},
    {"deal_id": "D-013-05", "rep_id": "REP-013", "close_date": "2026-03-08", "arr_usd": 81200.0},
    {"deal_id": "D-014-01", "rep_id": "REP-014", "close_date": "2026-01-13", "arr_usd": 21160.0},
    {"deal_id": "D-014-02", "rep_id": "REP-014", "close_date": "2026-02-07", "arr_usd": 21160.0},
    {"deal_id": "D-014-03", "rep_id": "REP-014", "close_date": "2026-03-02", "arr_usd": 21160.0},
    {"deal_id": "D-014-04", "rep_id": "REP-014", "close_date": "2026-03-19", "arr_usd": 21160.0},
    {"deal_id": "D-015-01", "rep_id": "REP-015", "close_date": "2026-01-19", "arr_usd": 22000.0},
    {"deal_id": "D-015-02", "rep_id": "REP-015", "close_date": "2026-02-12", "arr_usd": 8875.0},
    {"deal_id": "D-015-03", "rep_id": "REP-015", "close_date": "2026-03-15", "arr_usd": 8875.0},
    {"deal_id": "D-016-01", "rep_id": "REP-016", "close_date": "2026-01-06", "arr_usd": 43412.0},
    {"deal_id": "D-016-02", "rep_id": "REP-016", "close_date": "2026-01-23", "arr_usd": 43413.0},
    {"deal_id": "D-016-03", "rep_id": "REP-016", "close_date": "2026-02-14", "arr_usd": 43412.0},
    {"deal_id": "D-016-04", "rep_id": "REP-016", "close_date": "2026-03-11", "arr_usd": 43413.0},
    {"deal_id": "D-017-01", "rep_id": "REP-017", "close_date": "2026-01-21", "arr_usd": 26950.0},
    {"deal_id": "D-017-02", "rep_id": "REP-017", "close_date": "2026-02-16", "arr_usd": 26950.0},
    {"deal_id": "D-017-03", "rep_id": "REP-017", "close_date": "2026-03-05", "arr_usd": 26950.0},
    {"deal_id": "D-017-04", "rep_id": "REP-017", "close_date": "2026-03-24", "arr_usd": 26950.0},
    {"deal_id": "D-018-01", "rep_id": "REP-018", "close_date": "2026-01-26", "arr_usd": 11220.0},
    {"deal_id": "D-018-02", "rep_id": "REP-018", "close_date": "2026-02-20", "arr_usd": 11220.0},
    {"deal_id": "D-018-03", "rep_id": "REP-018", "close_date": "2026-03-16", "arr_usd": 11220.0},
    {"deal_id": "D-018-04", "rep_id": "REP-018", "close_date": "2026-03-28", "arr_usd": 11220.0},
    {"deal_id": "D-019-01", "rep_id": "REP-019", "close_date": "2026-01-05", "arr_usd": 61380.0},
    {"deal_id": "D-019-02", "rep_id": "REP-019", "close_date": "2026-01-29", "arr_usd": 61380.0},
    {"deal_id": "D-019-03", "rep_id": "REP-019", "close_date": "2026-02-23", "arr_usd": 61380.0},
    {"deal_id": "D-019-04", "rep_id": "REP-019", "close_date": "2026-03-17", "arr_usd": 61380.0},
    {"deal_id": "D-019-05", "rep_id": "REP-019", "close_date": "2026-03-30", "arr_usd": 61380.0},
    {"deal_id": "D-020-01", "rep_id": "REP-020", "close_date": "2026-01-31", "arr_usd": 12000.0},
    {"deal_id": "D-020-02", "rep_id": "REP-020", "close_date": "2026-02-28", "arr_usd": 12000.0},
    {"deal_id": "D-020-03", "rep_id": "REP-020", "close_date": "2026-03-27", "arr_usd": 12000.0},
]

# Draft with intentional discrepancies (>=3): REP-001, REP-003, REP-007, REP-012
DRAFT_COMMISSIONS = {
    "REP-001": 17200.00,
    "REP-002": 0.00,
    "REP-003": 50500.00,
    "REP-004": 7560.00,
    "REP-005": 37260.00,
    "REP-006": 0.00,
    "REP-007": 18800.00,
    "REP-008": 6688.00,
    "REP-009": 4016.00,
    "REP-010": 0.00,
    "REP-011": 12810.00,
    "REP-012": 11500.00,
    "REP-013": 60720.00,
    "REP-014": 6771.20,
    "REP-015": 0.00,
    "REP-016": 17365.00,
    "REP-017": 12936.00,
    "REP-018": 3590.40,
    "REP-019": 46035.00,
    "REP-020": 0.00,
}


def excel_round(value, ndigits=0):
    """Excel ROUND semantics (round half away from zero)."""
    ndigits = int(ndigits)
    if ndigits >= 0:
        factor = 10 ** ndigits
        return math.copysign(math.floor(abs(value) * factor + 0.5), value) / factor
    factor = 10 ** (-ndigits)
    return math.copysign(math.floor(abs(value) / factor + 0.5), value) * factor


def compute_carryover(base_quota, carry_rate, formula):
    raw = base_quota * carry_rate
    if formula == "MAX":
        return float(max(raw, 5000))
    return float(excel_round(raw, -3))


def tier_and_rate(att):
    if att < 0.80:
        return "Below Threshold", 0.00
    if att < 1.00:
        return "Base", 0.08
    if att < 1.20:
        return "Accelerator 1", 0.10
    if att < 1.50:
        return "Accelerator 2", 0.12
    return "Accelerator 3", 0.15


def flag_for(delta):
    if delta > 0.01:
        return "OVERPAID"
    if delta < -0.01:
        return "UNDERPAID"
    return "CORRECT"


AUDIT_HEADERS = [
    "rep_id", "base_quota_usd", "carryover_quota_usd", "total_quota_usd",
    "bookings_arr_usd", "attainment_pct", "commission_tier",
    "commission_usd", "draft_commission_usd", "delta_usd", "flag",
    "reviewer_note",
]

FIRST_DATA_ROW = 5
LAST_DATA_ROW = 24
STALE_LAST_DATA_ROW = 12

EDITABLE_AUDIT_COLS = {
    "base_quota_usd": 2,
    "carryover_quota_usd": 3,
    "bookings_arr_usd": 5,
    "commission_usd": 8,
    "draft_commission_usd": 9,
}

FORMULA_AUDIT_COLS = {
    "total_quota_usd": 4,
    "attainment_pct": 6,
    "commission_tier": 7,
    "delta_usd": 10,
    "flag": 11,
}

TABLE_NAME = "CommissionAuditTable"
EXPECTED_TABLE_RANGE = "A4:L24"
STALE_TABLE_RANGE = "A4:L12"

STALE_DEFINED_NAMES = {
    "AuditDataRange": "'Commission Audit'!$A$5:$L$12",
    "TierEntryRange": "'Commission Audit'!$G$5:$G$12",
    "FlagEntryRange": "'Commission Audit'!$K$5:$K$12",
}

EXPECTED_DEFINED_NAMES = {
    "AuditDataRange": "'Commission Audit'!$A$5:$L$24",
    "TierEntryRange": "'Commission Audit'!$G$5:$G$24",
    "FlagEntryRange": "'Commission Audit'!$K$5:$K$24",
}

STALE_CONDITIONAL_FORMAT_RANGES = {
    "attainment": "F5:F12",
    "delta": "J5:J12",
    "flag": "K5:K12",
}

EXPECTED_CONDITIONAL_FORMAT_RANGES = {
    "attainment": "F5:F24",
    "delta": "J5:J24",
    "flag": "K5:K24",
}

CONTROL_METADATA = {
    "editable_columns": "B,C,E,H,I",
    "formula_columns": "D,F,G,J,K",
    "preserve_columns": "A,L",
    "first_data_row": FIRST_DATA_ROW,
    "last_data_row": LAST_DATA_ROW,
    "active_range": "A4:L24",
    "validation_tier_range": "G5:G24",
    "validation_flag_range": "K5:K24",
    "summary_formula_range": f"{FIRST_DATA_ROW}:{LAST_DATA_ROW}",
    "table_name": TABLE_NAME,
    "expected_table_range": EXPECTED_TABLE_RANGE,
    "defined_names": "AuditDataRange,TierEntryRange,FlagEntryRange",
    "expected_AuditDataRange": EXPECTED_DEFINED_NAMES["AuditDataRange"],
    "expected_TierEntryRange": EXPECTED_DEFINED_NAMES["TierEntryRange"],
    "expected_FlagEntryRange": EXPECTED_DEFINED_NAMES["FlagEntryRange"],
    "conditional_format_ranges": "F5:F24,J5:J24,K5:K24",
}

HELPER_ROW_FORMULAS = {
    "B25": "=SUM(B5:B24)",
    "C25": "=SUM(C5:C24)",
    "E25": "=SUM(E5:E24)",
    "H25": "=SUM(H5:H24)",
    "J25": "=SUM(J5:J24)",
}

STALE_HELPER_ROW_FORMULAS = {
    "B25": "=SUM(B5:B12)",
    "C25": "=SUM(C5:C12)",
    "E25": "=SUM(E5:E12)",
    "H25": "=SUM(H5:H12)",
    "J25": "=SUM(J5:J12)",
}

REVIEWER_NOTES = {
    rep["rep_id"]: "Manual review: draft value from payroll export"
    for rep in REPS
}

SUMMARY_FORMULAS = {
    "B3": "=COUNTA('Commission Audit'!A5:A24)",
    "B4": "=COUNTIF('Commission Audit'!K5:K24,\"<>CORRECT\")",
    "B5": "=SUMIF('Commission Audit'!K5:K24,\"OVERPAID\",'Commission Audit'!J5:J24)",
    "B6": "=ABS(SUMIF('Commission Audit'!K5:K24,\"UNDERPAID\",'Commission Audit'!J5:J24))",
    "B7": "=IFERROR(INDEX('Commission Audit'!A5:A24,MATCH(MAX(ABS('Commission Audit'!J5:J24)),ABS('Commission Audit'!J5:J24),0)),\"\")",
    "B9": "=COUNTIF('Commission Audit'!G5:G24,\"Below Threshold\")",
    "B10": "=COUNTIF('Commission Audit'!G5:G24,\"Base\")",
    "B11": "=COUNTIF('Commission Audit'!G5:G24,\"Accelerator 1\")",
    "B12": "=COUNTIF('Commission Audit'!G5:G24,\"Accelerator 2\")",
    "B13": "=COUNTIF('Commission Audit'!G5:G24,\"Accelerator 3\")",
}

STALE_SUMMARY_FORMULAS = {
    "B3": "=COUNTA('Commission Audit'!A5:A12)",
    "B4": "=COUNTIF('Commission Audit'!K5:K12,\"<>CORRECT\")",
    "B5": "=SUMIF('Commission Audit'!K5:K12,\"OVERPAID\",'Commission Audit'!J5:J12)",
    "B6": "=ABS(SUMIF('Commission Audit'!K5:K12,\"UNDERPAID\",'Commission Audit'!J5:J12))",
    "B7": "=IFERROR(INDEX('Commission Audit'!A5:A12,MATCH(MAX(ABS('Commission Audit'!J5:J12)),ABS('Commission Audit'!J5:J12),0)),\"\")",
    "B9": "=COUNTIF('Commission Audit'!G5:G12,\"Below Threshold\")",
    "B10": "=COUNTIF('Commission Audit'!G5:G12,\"Base\")",
    "B11": "=COUNTIF('Commission Audit'!G5:G12,\"Accelerator 1\")",
    "B12": "=COUNTIF('Commission Audit'!G5:G12,\"Accelerator 2\")",
    "B13": "=COUNTIF('Commission Audit'!G5:G12,\"Accelerator 3\")",
}

# Stale editable-cell values on Commission Audit rows (before agent repair).
# Formula columns D, F, J are never written as literals in the draft.
STALE_AUDIT_ROWS = {
    "REP-001": {
        "base_quota_usd": 145000, "carryover_quota_usd": 0,
        "bookings_arr_usd": 168000, "commission_usd": 17200,
    },
    "REP-003": {
        "base_quota_usd": 180000, "carryover_quota_usd": 0,
        "bookings_arr_usd": 333300, "commission_usd": 50500,
    },
    "REP-007": {
        "base_quota_usd": 160000, "carryover_quota_usd": 0,
        "bookings_arr_usd": 193600, "commission_usd": 18800,
    },
    "REP-012": {
        "base_quota_usd": 65000, "carryover_quota_usd": 4000,
        "bookings_arr_usd": 91000, "commission_usd": 11500,
    },
}


def audit_row_formulas(row: int) -> dict[str, str]:
    return {
        "D": f"=B{row}+C{row}",
        "F": f"=E{row}/D{row}",
        "G": (
            f'=IF(F{row}<0.8,"Below Threshold",IF(F{row}<1,"Base",'
            f'IF(F{row}<1.2,"Accelerator 1",IF(F{row}<1.5,"Accelerator 2","Accelerator 3"))))'
        ),
        "J": f"=I{row}-H{row}",
        "K": (
            f'=IF(J{row}>0.01,"OVERPAID",IF(J{row}<-0.01,"UNDERPAID","CORRECT"))'
        ),
    }


# Stale formula-column literals for rows 13-24 (prior 8-rep template not extended).
STALE_FORMULA_COL_LITERALS = {
    13: {"D": 59000, "F": 0.851, "G": "Base", "J": 0, "K": "CORRECT"},
    14: {"D": 92000, "F": 0.92, "G": "Base", "J": 0, "K": "CORRECT"},
    15: {"D": 55000, "F": 0.723, "G": "Below Threshold", "J": 0, "K": "CORRECT"},
    16: {"D": 151000, "F": 1.15, "G": "Accelerator 1", "J": 0, "K": "CORRECT"},
    17: {"D": 77000, "F": 1.4, "G": "Accelerator 2", "J": 0, "K": "CORRECT"},
    18: {"D": 51000, "F": 0.88, "G": "Base", "J": 0, "K": "CORRECT"},
    19: {"D": 198000, "F": 1.55, "G": "Accelerator 3", "J": 0, "K": "CORRECT"},
    20: {"D": 60000, "F": 0.6, "G": "Below Threshold", "J": 0, "K": "CORRECT"},
    21: {"D": None, "F": None, "G": None, "J": None, "K": None},
    22: {"D": 82000, "F": 0.7, "G": "Below Threshold", "J": 0, "K": "CORRECT"},
    23: {"D": 122000, "F": 1.05, "G": "Accelerator 1", "J": 0, "K": "CORRECT"},
    24: {"D": 70000, "F": 1.479, "G": "Accelerator 2", "J": -920, "K": "UNDERPAID"},
}


def add_stale_conditional_formatting(ws):
    """Stale CF ranges stop at row 12; agent must extend to row 24."""
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

    ws.conditional_formatting.add(
        STALE_CONDITIONAL_FORMAT_RANGES["attainment"],
        ColorScaleRule(
            start_type="num", start_value=0, start_color="F8696B",
            mid_type="num", mid_value=1, mid_color="FFEB84",
            end_type="num", end_value=1.5, end_color="63BE7B",
        ),
    )
    ws.conditional_formatting.add(
        STALE_CONDITIONAL_FORMAT_RANGES["delta"],
        CellIsRule(operator="greaterThan", formula=["0"], fill=green_fill),
    )
    ws.conditional_formatting.add(
        STALE_CONDITIONAL_FORMAT_RANGES["delta"],
        CellIsRule(operator="lessThan", formula=["0"], fill=red_fill),
    )
    ws.conditional_formatting.add(
        STALE_CONDITIONAL_FORMAT_RANGES["flag"],
        CellIsRule(operator="equal", formula=['"OVERPAID"'], fill=red_fill),
    )
    ws.conditional_formatting.add(
        STALE_CONDITIONAL_FORMAT_RANGES["flag"],
        CellIsRule(operator="equal", formula=['"UNDERPAID"'], fill=yellow_fill),
    )
    ws.conditional_formatting.add(
        STALE_CONDITIONAL_FORMAT_RANGES["flag"],
        CellIsRule(operator="equal", formula=['"CORRECT"'], fill=green_fill),
    )


def add_stale_defined_names(wb):
    for name, attr_text in STALE_DEFINED_NAMES.items():
        wb.defined_names.add(DefinedName(name=name, attr_text=attr_text))


def build_draft_commission_audit_workbook(path):
    """Create draft audit workbook artifact the agent must repair in place."""
    wb = Workbook()

    cover = wb.active
    cover.title = "Cover"
    cover["A1"] = "Q1 2026 Commission Audit Workbook"
    cover["A1"].font = Font(bold=True, size=14)
    cover.column_dimensions["A"].width = 48
    cover["A2"] = "Draft workbook generated by Revenue Operations"
    cover["A4"] = "Instructions: Read the Control sheet before editing Commission Audit."
    cover["A5"] = "Do not delete, rename, or reorder sheets."

    audit = wb.create_sheet("Commission Audit")
    audit["A1"] = "Q1 2026 Commission Audit"
    audit["A2"] = "Source workbook generated before final close"
    for col, header in enumerate(AUDIT_HEADERS, start=1):
        cell = audit.cell(row=4, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(
            start_color="DDDDDD", end_color="DDDDDD", fill_type="solid"
        )

    for i, rep in enumerate(REPS):
        row = i + 5
        rep_id = rep["rep_id"]
        audit.cell(row=row, column=1, value=rep_id)
        stale = STALE_AUDIT_ROWS.get(rep_id, {})
        for field, col_idx in EDITABLE_AUDIT_COLS.items():
            if field == "draft_commission_usd":
                audit.cell(row=row, column=col_idx, value=DRAFT_COMMISSIONS[rep_id])
            elif field in stale:
                audit.cell(row=row, column=col_idx, value=stale[field])
        if row <= STALE_LAST_DATA_ROW:
            for col_letter, formula in audit_row_formulas(row).items():
                audit[f"{col_letter}{row}"] = formula
        else:
            stale_literals = STALE_FORMULA_COL_LITERALS.get(row, {})
            for col_letter, value in stale_literals.items():
                if value is not None:
                    audit[f"{col_letter}{row}"] = value
        audit.cell(row=row, column=12, value=REVIEWER_NOTES[rep_id])

    for cell_ref, formula in STALE_HELPER_ROW_FORMULAS.items():
        audit[cell_ref] = formula
    audit.row_dimensions[25].hidden = True

    audit.freeze_panes = "A5"
    audit.auto_filter.ref = "A4:L12"
    currency_fmt = "$#,##0.00"
    pct_fmt = "0.0000"
    for row in range(5, 25):
        for col in (2, 3, 4, 5, 8, 9, 10):
            audit.cell(row=row, column=col).number_format = currency_fmt
        audit.cell(row=row, column=6).number_format = pct_fmt

    tier_dv = DataValidation(
        type="list",
        formula1="Validation!$A$2:$A$6",
        allow_blank=True,
    )
    tier_dv.add("G5:G12")
    audit.add_data_validation(tier_dv)
    flag_dv = DataValidation(
        type="list",
        formula1="Validation!$C$2:$C$4",
        allow_blank=True,
    )
    flag_dv.add("K5:K12")
    audit.add_data_validation(flag_dv)

    audit_table = Table(displayName=TABLE_NAME, ref=STALE_TABLE_RANGE)
    audit_table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    audit.add_table(audit_table)
    add_stale_conditional_formatting(audit)
    add_stale_defined_names(wb)

    summary = wb.create_sheet("Summary")
    summary["A1"] = "Q1 Commission Audit Summary"
    summary["A3"], summary["B3"] = "total_reps_audited", STALE_SUMMARY_FORMULAS["B3"]
    summary["A4"], summary["B4"] = "reps_with_discrepancy", STALE_SUMMARY_FORMULAS["B4"]
    summary["A5"], summary["B5"] = "total_overpaid_usd", STALE_SUMMARY_FORMULAS["B5"]
    summary["A6"], summary["B6"] = "total_underpaid_usd", STALE_SUMMARY_FORMULAS["B6"]
    summary["A7"], summary["B7"] = "highest_delta_rep", STALE_SUMMARY_FORMULAS["B7"]
    summary["A9"], summary["B9"] = "Below Threshold", STALE_SUMMARY_FORMULAS["B9"]
    summary["A10"], summary["B10"] = "Base", STALE_SUMMARY_FORMULAS["B10"]
    summary["A11"], summary["B11"] = "Accelerator 1", STALE_SUMMARY_FORMULAS["B11"]
    summary["A12"], summary["B12"] = "Accelerator 2", STALE_SUMMARY_FORMULAS["B12"]
    summary["A13"], summary["B13"] = "Accelerator 3", STALE_SUMMARY_FORMULAS["B13"]

    control = wb.create_sheet("Control")
    control["A1"], control["B1"] = "control_name", "value"
    control["A2"], control["B2"] = "expected_row_count", 20
    control["A3"], control["B3"] = "workbook_version", "Q1-2026-v1"
    meta_start = 4
    for idx, (key, val) in enumerate(CONTROL_METADATA.items(), start=meta_start):
        control.cell(row=idx, column=1, value=key)
        control.cell(row=idx, column=2, value=val)
    tier_row = meta_start + len(CONTROL_METADATA)
    control.cell(row=tier_row, column=1, value="allowed_tiers")
    for i, tier in enumerate(
        ["Below Threshold", "Base", "Accelerator 1", "Accelerator 2", "Accelerator 3"],
        start=tier_row + 1,
    ):
        control.cell(row=i, column=1, value=tier)
    control.cell(row=tier_row, column=3, value="allowed_flags")
    for i, flag in enumerate(["OVERPAID", "UNDERPAID", "CORRECT"], start=tier_row + 1):
        control.cell(row=i, column=3, value=flag)
    control.sheet_state = "hidden"

    validation = wb.create_sheet("Validation")
    validation["A1"] = "tier_labels"
    for i, tier in enumerate(
        ["Below Threshold", "Base", "Accelerator 1", "Accelerator 2", "Accelerator 3"],
        start=2,
    ):
        validation.cell(row=i, column=1, value=tier)
    validation["C1"] = "flag_labels"
    for i, flag in enumerate(["OVERPAID", "UNDERPAID", "CORRECT"], start=2):
        validation.cell(row=i, column=3, value=flag)
    validation.sheet_state = "hidden"

    notes = wb.create_sheet("Notes")
    notes["A1"] = "Notes"
    notes["A3"] = "Do not delete hidden control or validation sheets."
    notes["A4"] = "Summary formulas are linked to the Commission Audit sheet."
    notes["A5"] = "Read Control sheet metadata before editing Commission Audit."

    wb.save(path)


def gen_commission_plan_pdf(path):
    c = canvas.Canvas(path, pagesize=A4)
    _, h = A4
    y = h - 72

    def line(txt, font="Helvetica", size=10, gap=14):
        nonlocal y
        c.setFont(font, size)
        c.drawString(72, y, txt)
        y -= gap

    line("Abundant Cloud Platform — Q1 2026 AE Commission Plan", "Helvetica-Bold", 13, 24)
    line("Effective period: January 1, 2026 – March 31, 2026", gap=13)
    line("Audience: Account Executive commission audit", gap=20)

    line("Plan overview", "Helvetica-Bold", 11, 16)
    line("Account Executives earn quarterly commission on New ARR booked during the", gap=13)
    line("period. Commission is determined by quota attainment against a total quota", gap=13)
    line("that includes any approved carryover allocation from the prior period.", gap=20)

    line("Quota components", "Helvetica-Bold", 11, 16)
    line("- Base quota: The rep's assigned Q1 2026 target.", gap=13)
    line("- Carryover quota: Additional quota credit allocated per rep. Approved", gap=13)
    line("  carryover amounts are recorded in quota_carryover.xlsx.", gap=13)
    line("Commission workbook formula values should be interpreted using", gap=13)
    line("standard Excel formula semantics.", gap=13)
    line("- Total quota: total_quota_usd = base_quota_usd + carryover_quota_usd", gap=13)
    line("Carryover amounts add to the quota denominator. They do not count as bookings.", gap=20)

    line("Attainment", "Helvetica-Bold", 11, 16)
    line("attainment_pct = bookings_arr_usd / total_quota_usd", gap=13)
    line("where bookings_arr_usd is the sum of arr_usd for all deals closed in Q1 2026", gap=13)
    line("for that rep.", gap=20)

    line("Commission tiers (ratchet schedule)", "Helvetica-Bold", 11, 16)
    line("Below Threshold: attainment < 80% — 0%", gap=13)
    line("Base: 80% <= attainment < 100% — 8%", gap=13)
    line("Accelerator 1: 100% <= attainment < 120% — 10%", gap=13)
    line("Accelerator 2: 120% <= attainment < 150% — 12%", gap=13)
    line("Accelerator 3: attainment >= 150% — 15%", gap=13)
    line("Boundary behavior: inclusive on the lower bound, exclusive on the upper bound.", gap=20)

    line("Accelerator stacking rule (non-marginal)", "Helvetica-Bold", 11, 16)
    line("The commission rate for the tier reached applies to ALL Q1 bookings for", gap=13)
    line("that rep — not marginally by tier band.", gap=20)

    line("Audit comparison", "Helvetica-Bold", 11, 16)
    line("Compare computed commission_usd against draft_commission_usd from", gap=13)
    line("draft_commission_statements.csv.", gap=13)
    line("delta_usd = draft_commission_usd - commission_usd", gap=13)
    line("OVERPAID if delta_usd > 0.01", gap=13)
    line("UNDERPAID if delta_usd < -0.01", gap=13)
    line("CORRECT if |delta_usd| <= 0.01", gap=20)

    c.save()


# 1. rep_quotas.csv — Trap A: currency-formatted strings
quota_rows = []
for r in REPS:
    quota_rows.append({
        "rep_id": r["rep_id"],
        "rep_name": r["rep_name"],
        "region": r["region"],
        "base_quota_usd": "${:,.2f}".format(r["base_quota"]),
    })
pd.DataFrame(quota_rows).to_csv(os.path.join(OUTPUT_DIR, "rep_quotas.csv"), index=False)

# 2. quota_carryover.xlsx — Cover, RateTable, Allocations, BaseRef
wb = Workbook()
cover = wb.active
cover.title = "Cover"
cover["A1"] = "Q1 2026 Quota Carryover Allocation"
cover["A3"] = "Prepared by Sales Operations — Internal Use Only"
cover["A5"] = "This workbook documents carryover credits applied to Q1 2026 quotas."
cover["A7"] = "Per-rep carryover amounts are computed in this workbook."

rate_ws = wb.create_sheet("RateTable")
rate_ws.cell(row=1, column=1, value="rep_id")
rate_ws.cell(row=1, column=2, value="carry_rate")
for i, r in enumerate(REPS, start=2):
    rate_ws.cell(row=i, column=1, value=r["rep_id"])
    rate_ws.cell(row=i, column=2, value=r["carry_rate"])

alloc_ws = wb.create_sheet("Allocations")
alloc_ws.cell(row=1, column=1, value="rep_id")
alloc_ws.cell(row=1, column=2, value="carryover_quota_usd")
vlookup_rate = "VLOOKUP(A{row},RateTable!$A$2:$B$21,2,FALSE)"
vlookup_base = "VLOOKUP(A{row},BaseRef!$A$2:$B$21,2,FALSE)"
for i, r in enumerate(REPS, start=2):
    alloc_ws.cell(row=i, column=1, value=r["rep_id"])
    rate_expr = vlookup_rate.format(row=i)
    base_expr = vlookup_base.format(row=i)
    product = f"{rate_expr}*{base_expr}"
    if r["formula"] == "MAX":
        alloc_ws.cell(row=i, column=2, value=f"=MAX({product},5000)")
    else:
        alloc_ws.cell(row=i, column=2, value=f"=ROUND({product},-3)")

base_ws = wb.create_sheet("BaseRef")
base_ws.cell(row=1, column=1, value="rep_id")
base_ws.cell(row=1, column=2, value="base_quota_usd")
for i, r in enumerate(REPS, start=2):
    base_ws.cell(row=i, column=1, value=r["rep_id"])
    base_ws.cell(row=i, column=2, value=r["base_quota"])

wb.save(os.path.join(OUTPUT_DIR, "quota_carryover.xlsx"))

# 3. bookings.jsonl — Trap C: mixed float and currency-string arr_usd
# String-valued deals appear after line 35 so early head() samples miss them.
def write_bookings_jsonl(path):
    string_deals = [d for d in DEALS if d["deal_id"] in STRING_ARR_DEALS]
    numeric_deals = [d for d in DEALS if d["deal_id"] not in STRING_ARR_DEALS]
    ordered = numeric_deals[:35] + string_deals + numeric_deals[35:]
    with open(path, "w", encoding="utf-8") as f:
        for d in ordered:
            row = dict(d)
            if row["deal_id"] in STRING_ARR_DEALS:
                row["arr_usd"] = STRING_ARR_DEALS[row["deal_id"]]
            f.write(json.dumps(row) + "\n")


write_bookings_jsonl(os.path.join(OUTPUT_DIR, "bookings.jsonl"))

# 4. commission_plan.pdf
gen_commission_plan_pdf(os.path.join(OUTPUT_DIR, "commission_plan.pdf"))

# 5. draft_commission_statements.csv
draft_rows = [{"rep_id": k, "draft_commission_usd": v} for k, v in DRAFT_COMMISSIONS.items()]
pd.DataFrame(draft_rows).to_csv(
    os.path.join(OUTPUT_DIR, "draft_commission_statements.csv"), index=False
)

# 6. draft_commission_audit.xlsx — workbook artifact to repair in place
build_draft_commission_audit_workbook(
    os.path.join(OUTPUT_DIR, "draft_commission_audit.xlsx")
)

# Sanity check: verify deal sums match targets
for rep_id, target in BOOKINGS_TARGET.items():
    total = 0.0
    for d in DEALS:
        if d["rep_id"] != rep_id:
            continue
        val = d["arr_usd"]
        if d["deal_id"] in STRING_ARR_DEALS:
            val = float(STRING_ARR_DEALS[d["deal_id"]].replace("$", "").replace(",", ""))
        total += val
    assert abs(total - target) < 0.01, f"{rep_id} bookings sum {total} != {target}"

print("Generated input files in", OUTPUT_DIR)
