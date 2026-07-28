"""approval_workbook.py — load and evaluate formula-backed exception approvals."""

from __future__ import annotations

import math
import re
from typing import Any

import openpyxl
import pandas as pd


APPROVAL_COLUMNS = [
    "approval_id",
    "account_id",
    "exception_type",
    "approval_status",
    "effective_date",
    "expiry_date",
    "approved_plan",
    "approved_price",
    "b_key",
]


def _norm_status(val: Any) -> str:
    if val is None:
        return ""
    return str(val).strip().lower()


_INACTIVE_STATUSES = frozenset({
    "expired", "cancelled", "superseded", "pending", "denied", "rejected",
})


def is_approval_status_active(status: Any) -> bool:
    s = _norm_status(status)
    if s in _INACTIVE_STATUSES:
        return False
    return s == "approved"


def is_approval_valid(row: dict, migration_date: str) -> bool:
    if not is_approval_status_active(row.get("approval_status")):
        return False
    migration_dt = pd.Timestamp(migration_date)
    try:
        effective = pd.Timestamp(row.get("effective_date"))
        if effective > migration_dt:
            return False
    except Exception:
        return False
    try:
        expiry = pd.Timestamp(row.get("expiry_date"))
        if expiry < migration_dt:
            return False
    except Exception:
        return False
    return True


def _sheet_table(wb, sheet_name: str) -> list[list[Any]]:
    ws = wb[sheet_name]
    return [list(r) for r in ws.iter_rows(values_only=True)]


def _build_vlookup_tables(wb) -> dict[str, list[list[Any]]]:
    skip = {"Cover", "Approvals"}
    tables: dict[str, list[list[Any]]] = {}
    for name in wb.sheetnames:
        if name in skip:
            continue
        rows = _sheet_table(wb, name)
        if rows:
            tables[name] = rows
    return tables


def _col_index(col_letter: str) -> int:
    col_letter = col_letter.upper()
    idx = 0
    for ch in col_letter:
        idx = idx * 26 + (ord(ch) - ord("A") + 1)
    return idx - 1


def _parse_range(table_rows: list[list[Any]], range_spec: str) -> list[list[Any]]:
    # LookupKeys!A:C or RateReference!A:C
    if "!" not in range_spec:
        return table_rows
    _, cols = range_spec.split("!", 1)
    if ":" not in cols:
        return table_rows
    start_col, end_col = cols.split(":", 1)
    start_i = _col_index(re.sub(r"\d+", "", start_col) or "A")
    end_i = _col_index(re.sub(r"\d+", "", end_col) or "Z")
    out = []
    for row in table_rows:
        out.append(list(row[start_i : end_i + 1]))
    return out


def _vlookup(key: Any, table: list[list[Any]], col_index: int) -> Any:
    if not table:
        return None
    header = table[0]
    data = table[1:]
    key_str = str(key).strip()
    key_col = 0
    for row in data:
        if row and len(row) > key_col and str(row[key_col]).strip() == key_str:
            idx = col_index - 1
            if 0 <= idx < len(row):
                return row[idx]
            return None
    return None


def _to_number(val: Any) -> float:
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip()
    if not s:
        return 0.0
    return float(s)


def _expr_depth(expr: str, idx: int) -> int:
    depth = 0
    for i in range(idx):
        ch = expr[i]
        if ch == "(":
            depth += 1
        elif ch == ")":
            depth -= 1
    return depth


def _split_top_level_commas(expr: str) -> list[str]:
    parts: list[str] = []
    depth = 0
    current: list[str] = []
    for ch in expr:
        if ch == "(":
            depth += 1
            current.append(ch)
        elif ch == ")":
            depth -= 1
            current.append(ch)
        elif ch == "," and depth == 0:
            parts.append("".join(current).strip())
            current = []
        else:
            current.append(ch)
    if current:
        parts.append("".join(current).strip())
    return parts


def _eval_expr(expr: str, ctx: dict[str, Any], tables: dict[str, list[list[Any]]]) -> Any:
    expr = expr.strip()
    if not expr:
        return None

    if expr.startswith('"') and expr.endswith('"'):
        return expr[1:-1]

    # Binary + and - at top level (lowest precedence)
    for i in range(len(expr) - 1, 0, -1):
        ch = expr[i]
        if ch in "+-" and _expr_depth(expr, i) == 0:
            left = _eval_expr(expr[:i], ctx, tables)
            right = _eval_expr(expr[i + 1 :], ctx, tables)
            if ch == "+":
                return _to_number(left) + _to_number(right)
            return _to_number(left) - _to_number(right)

    # Binary * and / at top level
    for i in range(len(expr) - 1, 0, -1):
        ch = expr[i]
        if ch in "*/" and _expr_depth(expr, i) == 0:
            left = _eval_expr(expr[:i], ctx, tables)
            right = _eval_expr(expr[i + 1 :], ctx, tables)
            if ch == "*":
                return _to_number(left) * _to_number(right)
            divisor = _to_number(right)
            return _to_number(left) / divisor if divisor else 0.0

    # IF(cond, a, b)
    if re.match(r"IF\(", expr, flags=re.IGNORECASE) and expr.endswith(")"):
        args = _split_top_level_commas(expr[3:-1])
        if len(args) == 3:
            cond_raw, a_raw, b_raw = args
            cond_val = _eval_expr(cond_raw, ctx, tables)
            branch = a_raw if bool(cond_val) else b_raw
            return _eval_expr(branch, ctx, tables)

    # MAX(a,b)
    if re.match(r"MAX\(", expr, flags=re.IGNORECASE) and expr.endswith(")"):
        args = _split_top_level_commas(expr[4:-1])
        if len(args) == 2:
            left = _eval_expr(args[0], ctx, tables)
            right = _eval_expr(args[1], ctx, tables)
            return max(_to_number(left), _to_number(right))

    # MIN(a,b)
    if re.match(r"MIN\(", expr, flags=re.IGNORECASE) and expr.endswith(")"):
        args = _split_top_level_commas(expr[4:-1])
        if len(args) == 2:
            left = _eval_expr(args[0], ctx, tables)
            right = _eval_expr(args[1], ctx, tables)
            return min(_to_number(left), _to_number(right))

    # ROUND(x, n)
    if re.match(r"ROUND\(", expr, flags=re.IGNORECASE) and expr.endswith(")"):
        args = _split_top_level_commas(expr[6:-1])
        if len(args) == 2:
            val = _to_number(_eval_expr(args[0], ctx, tables))
            digits = int(float(_eval_expr(args[1], ctx, tables)))
            factor = 10 ** digits
            return float(round(val * factor) / factor)

    # ROUNDUP(x, -2)
    if re.match(r"ROUNDUP\(", expr, flags=re.IGNORECASE) and expr.endswith(")"):
        args = _split_top_level_commas(expr[8:-1])
        if len(args) == 2:
            val = _to_number(_eval_expr(args[0], ctx, tables))
            digits = int(float(_eval_expr(args[1], ctx, tables)))
            if digits < 0:
                factor = 10 ** abs(digits)
                return float(math.ceil(val / factor) * factor)
            factor = 10 ** digits
            return float(math.ceil(val * factor) / factor)

    # VLOOKUP(key, Sheet!A:E, n, FALSE)
    if re.match(r"VLOOKUP\(", expr, flags=re.IGNORECASE) and expr.endswith(")"):
        args = _split_top_level_commas(expr[8:-1])
        if len(args) == 4:
            key = _eval_expr(args[0], ctx, tables)
            range_spec = args[1]
            col_index = int(args[2])
            sheet_name = range_spec.split("!", 1)[0]
            table_rows = tables.get(sheet_name, [])
            sliced = _parse_range(table_rows, range_spec)
            return _vlookup(key, sliced, col_index)

    # Cell reference like A7 or I9 (same-row basis / id lookups in formulas)
    m = re.fullmatch(r"([A-Z]+)(\d+)", expr, flags=re.IGNORECASE)
    if m:
        col_letter, row_idx = m.group(1).upper(), int(m.group(2))
        col_i = _col_index(col_letter)
        headers = ctx.get("headers", [])
        row_dict = ctx.get("row_values", {}).get(row_idx, {})
        if col_i < len(headers):
            return row_dict.get(headers[col_i])
        if col_letter == "A":
            return row_dict.get("approval_id")
        return None

    # Numeric literal
    try:
        return float(expr)
    except ValueError:
        pass

    # Comparison / equality for IF conditions
    eq_idx = None
    depth = 0
    for i, ch in enumerate(expr):
        if ch == "(":
            depth += 1
        elif ch == ")":
            depth -= 1
        elif ch == "=" and depth == 0:
            eq_idx = i
            break
    if eq_idx is not None:
        left = _eval_expr(expr[:eq_idx].strip(), ctx, tables)
        right_raw = expr[eq_idx + 1 :].strip()
        if right_raw.startswith('"') and right_raw.endswith('"'):
            right = right_raw[1:-1]
        else:
            right = _eval_expr(right_raw, ctx, tables)
        return str(left).strip().lower() == str(right).strip().lower()

    raise ValueError(f"Unsupported expression: {expr}")


def evaluate_cell_formula(formula: Any, ctx: dict[str, Any], tables: dict[str, list[list[Any]]]) -> Any:
    if formula is None:
        return None
    if not isinstance(formula, str):
        return formula
    val = formula.strip()
    if not val.startswith("="):
        return formula
    expr = val[1:].strip()
    return _eval_expr(expr, ctx, tables)


def load_exception_approvals(data_dir: str) -> pd.DataFrame:
    path = f"{data_dir}/exception_approvals.xlsx"
    wb = openpyxl.load_workbook(path, data_only=False)
    ws = wb["Approvals"]
    rows = list(ws.iter_rows(values_only=True))

    header_idx = None
    for i, row in enumerate(rows):
        cells = [str(c).strip().lower() if c is not None else "" for c in row]
        if "approval_id" in cells:
            header_idx = i
            break
    if header_idx is None:
        return pd.DataFrame(columns=APPROVAL_COLUMNS)

    headers = [str(c).strip() if c is not None else "" for c in rows[header_idx]]
    tables = _build_vlookup_tables(wb)

    records = []
    row_values: dict[int, dict[str, Any]] = {}
    for i, row in enumerate(rows[header_idx + 1 :], start=header_idx + 2):
        if not row or all(c is None for c in row):
            continue
        base = {headers[j]: row[j] for j in range(min(len(headers), len(row)))}
        if not base.get("approval_id"):
            continue
        row_values[i] = base
        ctx = {"row_values": row_values, "current_row": i, "headers": headers}

        approved_plan = base.get("approved_plan")
        approved_price = base.get("approved_price")
        if isinstance(approved_plan, str) and approved_plan.startswith("="):
            approved_plan = evaluate_cell_formula(approved_plan, ctx, tables)
        if isinstance(approved_price, str) and approved_price.startswith("="):
            approved_price = evaluate_cell_formula(approved_price, ctx, tables)

        row_values[i]["approved_plan"] = approved_plan
        row_values[i]["approved_price"] = approved_price

        approval_type = base.get("approval_type") or base.get("exception_type")
        records.append(
            {
                "approval_id": str(base.get("approval_id")).strip(),
                "account_id": str(base.get("account_id")).strip(),
                "exception_type": str(approval_type).strip(),
                "approval_status": base.get("approval_status"),
                "effective_date": base.get("effective_date"),
                "expiry_date": base.get("expiry_date"),
                "approved_plan": approved_plan,
                "approved_price": approved_price,
                "b_key": base.get("b_key") or base.get("approval_basis"),
            }
        )

    df = pd.DataFrame(records)
    for col in ("effective_date", "expiry_date"):
        if col in df.columns:
            df[col] = pd.to_datetime(df[col], errors="coerce")
    return df


def valid_approvals(df: pd.DataFrame, migration_date: str) -> pd.DataFrame:
    if df is None or df.empty:
        return df
    mask = df.apply(lambda r: is_approval_valid(r.to_dict(), migration_date), axis=1)
    return df[mask].copy()
