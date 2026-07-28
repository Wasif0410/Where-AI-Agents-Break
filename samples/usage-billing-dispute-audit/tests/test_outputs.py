"""
Verifier for the usage-billing-dispute-audit task.

All expected values are derived deterministically from build_inputs.py.
Tests are fully deterministic and do not use LLM-as-judge.

Cascading failure design:
  If PDF extraction fails (wrong billed amounts) → Reconciliation deltas wrong → Exceptions wrong → Summary wrong.
  If pricing rules not read from addendum (wrong tier/discount) → Expected charges wrong → everything downstream wrong.
  If legacy workbook formulas not evaluated → billable/tier2 wrong → expected charges wrong.
  If rounding not applied → Expected charges slightly off → delta wrong.
  If tax not separated (acct_003 Feb) → Expected wrong, exception type wrong.
  If fx_rates.csv ignored (acct_002 EUR) → Billed USD wrong → deltas wrong.
  If WRONG_TIER not classified (acct_003 Mar) → Exception/summary wrong.
"""

from __future__ import annotations
from pathlib import Path
from openpyxl import load_workbook

OUTPUT_PATH = Path("/root/out/billing_dispute_audit.xlsx")
TOLERANCE   = 0.01

REQUIRED_SHEETS = [
    "Extracted Invoice Lines",
    "Expected Usage Charges",
    "Billing Reconciliation",
    "Dispute Exceptions",
    "Executive Summary",
]


def load_wb():
    assert OUTPUT_PATH.exists(), f"Missing: {OUTPUT_PATH}"
    return load_workbook(OUTPUT_PATH, data_only=True)


def load_wb_formulas():
    assert OUTPUT_PATH.exists(), f"Missing: {OUTPUT_PATH}"
    return load_workbook(OUTPUT_PATH, data_only=False)


def sheet_rows(ws, header_row=1):
    headers = {}
    for idx, cell in enumerate(ws[header_row], start=1):
        if cell.value is not None:
            headers[str(cell.value).strip()] = idx
    rows = []
    for row_cells in ws.iter_rows(min_row=header_row + 1):
        vals = tuple(c.value for c in row_cells)
        if not any(v is not None for v in vals):
            break
        rows.append(vals)
    return headers, rows


# ─────────────────────────────────────────────────────────────────────────────
# SECTION 1: STRUCTURE
# ─────────────────────────────────────────────────────────────────────────────

class TestStructure:
    def test_workbook_exists(self):
        assert OUTPUT_PATH.exists(), f"Missing: {OUTPUT_PATH}"

    def test_required_sheets_present(self):
        wb = load_wb()
        for name in REQUIRED_SHEETS:
            assert name in wb.sheetnames, f"Sheet '{name}' missing"

    def test_sheet_order(self):
        wb = load_wb()
        actual = [s for s in wb.sheetnames if s in REQUIRED_SHEETS]
        assert actual == REQUIRED_SHEETS, f"Wrong sheet order: {actual}"


# ─────────────────────────────────────────────────────────────────────────────
# SECTION 2: EXTRACTED INVOICE LINES — tests PDF parsing of 3 different formats
# ─────────────────────────────────────────────────────────────────────────────

class TestExtractedInvoiceLines:
    REQUIRED_COLS = {"source_file", "account_id", "invoice_month", "line_description",
                     "quantity", "unit_price", "line_amount", "line_type"}

    def _get_lines(self):
        wb = load_wb()
        ws = wb["Extracted Invoice Lines"]
        return sheet_rows(ws)

    def test_columns_exact(self):
        headers, _ = self._get_lines()
        missing = self.REQUIRED_COLS - set(headers)
        extra   = set(headers) - self.REQUIRED_COLS
        assert not missing, f"Missing columns: {missing}"
        assert not extra,   f"Extra columns: {extra}"

    def test_acct_001_feb_overage_extracted(self):
        """invoice_acct_001.pdf is a pipe-delimited table.
        Feb overage: 46,000 calls at $0.002 = $92.00. Correct, no dispute.
        """
        headers, rows = self._get_lines()
        acct_col   = headers["account_id"] - 1
        month_col  = headers["invoice_month"] - 1
        type_col   = headers["line_type"] - 1
        amount_col = headers["line_amount"] - 1

        feb_overage = [r for r in rows
                       if str(r[acct_col]).strip() == "acct_001"
                       and str(r[month_col]).strip() == "2026-02"
                       and str(r[type_col]).strip() == "overage"]
        assert feb_overage, "acct_001 Feb 2026 overage line missing from Extracted Invoice Lines"
        v = float(feb_overage[0][amount_col])
        assert abs(v - 92.00) <= TOLERANCE, (
            f"acct_001 Feb overage: expected 92.00 from pipe-table PDF, got {v}"
        )

    def test_acct_002_jan_overage_extracted(self):
        """invoice_acct_002.pdf uses section-based prose format (EUR).
        Jan overage billed as EUR 150.00 (discount missing on invoice).
        """
        headers, rows = self._get_lines()
        acct_col   = headers["account_id"] - 1
        month_col  = headers["invoice_month"] - 1
        type_col   = headers["line_type"] - 1
        amount_col = headers["line_amount"] - 1

        jan_overage = [r for r in rows
                       if str(r[acct_col]).strip() == "acct_002"
                       and str(r[month_col]).strip() == "2026-01"
                       and str(r[type_col]).strip() == "overage"]
        assert jan_overage, (
            "acct_002 Jan 2026 overage line missing. "
            "invoice_acct_002.pdf uses section-based prose — parse each monthly section."
        )
        v = float(jan_overage[0][amount_col])
        assert abs(v - 150.00) <= TOLERANCE, (
            f"acct_002 Jan billed overage: expected 150.00 EUR (as billed), got {v}"
        )

    def test_acct_003_mar_overage_extracted(self):
        """acct_003 Mar overage billed at flat Tier 1 rate."""
        headers, rows = self._get_lines()
        acct_col   = headers["account_id"] - 1
        month_col  = headers["invoice_month"] - 1
        type_col   = headers["line_type"] - 1
        amount_col = headers["line_amount"] - 1

        mar_overage = [r for r in rows
                       if str(r[acct_col]).strip() == "acct_003"
                       and str(r[month_col]).strip() == "2026-03"
                       and str(r[type_col]).strip() == "overage"]
        assert mar_overage, "acct_003 Mar 2026 overage line missing"
        v = float(mar_overage[0][amount_col])
        assert abs(v - 240.00) <= TOLERANCE, (
            f"acct_003 Mar overage: expected 240.00, got {v}"
        )

    def test_acct_003_feb_overage_and_tax_extracted(self):
        """invoice_acct_003.pdf uses inline prose with embedded tax."""
        headers, rows = self._get_lines()
        acct_col   = headers["account_id"] - 1
        month_col  = headers["invoice_month"] - 1
        type_col   = headers["line_type"] - 1
        amount_col = headers["line_amount"] - 1

        feb_lines = [r for r in rows
                     if str(r[acct_col]).strip() == "acct_003"
                     and str(r[month_col]).strip() == "2026-02"]
        assert feb_lines, (
            "acct_003 Feb 2026 lines missing. "
            "invoice_acct_003.pdf uses inline prose — extract from text like "
            "'32,000 calls at $0.002/call = $64.00 (incl. tax $10.24)'."
        )
        feb_amounts = [float(r[amount_col]) for r in feb_lines if r[amount_col] is not None]
        has_total  = any(abs(a - 74.24) <= TOLERANCE for a in feb_amounts)
        has_split  = (any(abs(a - 64.00) <= TOLERANCE for a in feb_amounts) and
                      any(abs(a - 10.24) <= TOLERANCE for a in feb_amounts))
        assert has_total or has_split, (
            f"acct_003 Feb extraction: expected either 74.24 (total) or 64.00+10.24 (split). "
            f"Got amounts: {feb_amounts}"
        )


# ─────────────────────────────────────────────────────────────────────────────
# SECTION 3: EXPECTED USAGE CHARGES — tests pricing rule application
# ─────────────────────────────────────────────────────────────────────────────

class TestExpectedUsageCharges:
    REQUIRED_COLS = {"account_id", "service_month", "actual_calls", "included_calls",
                     "billable_calls", "expected_overage_usd", "discount_applied",
                     "pricing_rule_source"}

    def _get_rows(self):
        wb = load_wb()
        ws = wb["Expected Usage Charges"]
        return sheet_rows(ws)

    def test_columns_present(self):
        headers, _ = self._get_rows()
        missing = self.REQUIRED_COLS - set(headers)
        assert not missing, f"Missing columns in Expected Usage Charges: {missing}"

    def test_acct_001_feb_billable_from_legacy_workbook(self):
        """acct_001 Feb: metered 145,000 calls but legacy workbook billable_calls = 46,000.
        Naive usage-only derivation yields 45,000 billable and $90 — wrong.
        """
        headers, rows = self._get_rows()
        acct_col = headers["account_id"] - 1
        month_col = headers["service_month"] - 1
        bill_col = headers["billable_calls"] - 1
        exp_col  = headers["expected_overage_usd"] - 1

        row = next((r for r in rows
                    if str(r[acct_col]).strip() == "acct_001"
                    and str(r[month_col]).strip() == "2026-02"), None)
        assert row is not None, "acct_001 2026-02 missing from Expected Usage Charges"
        billable = int(row[bill_col])
        charge = float(row[exp_col])
        assert billable == 46000, (
            f"acct_001 Feb billable_calls: expected 46000 (legacy workbook), got {billable}. "
            f"45000 means legacy_call_adjustments.xlsx was not evaluated."
        )
        assert abs(charge - 92.00) <= TOLERANCE, (
            f"acct_001 Feb expected_overage_usd: expected 92.00, got {charge}"
        )

    def test_acct_002_jan_expected_with_discount(self):
        """acct_002 Jan: 175,000 calls. 75,000 billable. 8% Enterprise Legacy discount → $138.00."""
        headers, rows = self._get_rows()
        acct_col  = headers["account_id"] - 1
        month_col = headers["service_month"] - 1
        exp_col   = headers["expected_overage_usd"] - 1

        row = next((r for r in rows
                    if str(r[acct_col]).strip() == "acct_002"
                    and str(r[month_col]).strip() == "2026-01"), None)
        assert row is not None, "acct_002 2026-01 missing from Expected Usage Charges"
        v = float(row[exp_col])
        assert abs(v - 138.00) <= TOLERANCE, (
            f"acct_002 Jan expected_overage_usd: expected 138.00 (150×0.92, Enterprise Legacy discount), "
            f"got {v}. Value 150.00 means discount from pricing_addendum.pdf was not applied."
        )

    def test_acct_002_mar_expected_tiered_with_legacy_tier2(self):
        """acct_002 Mar: 350,000 calls. 250,000 billable. tier2_calls = 150,000 from legacy workbook.
        Raw total: $425.00. After 8% discount: $391.00.
        """
        headers, rows = self._get_rows()
        acct_col  = headers["account_id"] - 1
        month_col = headers["service_month"] - 1
        t2_col    = headers.get("tier2_calls", 0) - 1
        exp_col   = headers["expected_overage_usd"] - 1

        row = next((r for r in rows
                    if str(r[acct_col]).strip() == "acct_002"
                    and str(r[month_col]).strip() == "2026-03"), None)
        assert row is not None, "acct_002 2026-03 missing from Expected Usage Charges"
        if t2_col >= 0:
            t2 = int(row[t2_col])
            assert t2 == 150000, (
                f"acct_002 Mar tier2_calls: expected 150000 (legacy MAX formula), got {t2}. "
                f"NaN/zero means legacy_call_adjustments.xlsx was not evaluated."
            )
        v = float(row[exp_col])
        assert abs(v - 391.00) <= TOLERANCE, (
            f"acct_002 Mar expected_overage_usd: expected 391.00 (T1+T2 tiered × 0.92 discount), "
            f"got {v}. 425.00=no discount, 460.00=wrong tier, 184.00=tier2 formula not evaluated"
        )

    def test_acct_003_feb_expected_excludes_tax(self):
        """acct_003 Feb: 132,000 calls. 32,000 billable. $64.00. Tax NOT included in expected."""
        headers, rows = self._get_rows()
        acct_col  = headers["account_id"] - 1
        month_col = headers["service_month"] - 1
        exp_col   = headers["expected_overage_usd"] - 1

        row = next((r for r in rows
                    if str(r[acct_col]).strip() == "acct_003"
                    and str(r[month_col]).strip() == "2026-02"), None)
        assert row is not None, "acct_003 2026-02 missing from Expected Usage Charges"
        v = float(row[exp_col])
        assert abs(v - 64.00) <= TOLERANCE, (
            f"acct_003 Feb expected_overage_usd: expected 64.00 (tax excluded), got {v}. "
            f"74.24 means tax was included in expected — it must be excluded per pricing addendum."
        )

    def test_acct_003_mar_expected_tiered(self):
        """acct_003 Mar: 220,000 calls. 120,000 billable. Tiered expected $230.00."""
        headers, rows = self._get_rows()
        acct_col  = headers["account_id"] - 1
        month_col = headers["service_month"] - 1
        exp_col   = headers["expected_overage_usd"] - 1
        t2_col    = headers.get("tier2_calls", 0) - 1

        row = next((r for r in rows
                    if str(r[acct_col]).strip() == "acct_003"
                    and str(r[month_col]).strip() == "2026-03"), None)
        assert row is not None, "acct_003 2026-03 missing from Expected Usage Charges"
        if t2_col >= 0:
            assert int(row[t2_col]) == 20000, (
                f"acct_003 Mar tier2_calls: expected 20000, got {row[t2_col]}"
            )
        v = float(row[exp_col])
        assert abs(v - 230.00) <= TOLERANCE, (
            f"acct_003 Mar expected_overage_usd: expected 230.00 (tiered), got {v}"
        )

    def test_pricing_rule_source_present(self):
        """Expected Usage Charges rows must cite the pricing source document."""
        headers, rows = self._get_rows()
        src_col = headers.get("pricing_rule_source", 0) - 1
        if src_col < 0:
            return

        missing = [str(r[0]) for r in rows
                   if r[src_col] is None or str(r[src_col]).strip() == ""]
        assert not missing, (
            f"These Expected Usage Charges rows have no pricing_rule_source: {missing}. "
            f"Every row must cite pricing_addendum.pdf."
        )


# ─────────────────────────────────────────────────────────────────────────────
# SECTION 4: BILLING RECONCILIATION — tests delta computation
# ─────────────────────────────────────────────────────────────────────────────

class TestBillingReconciliation:
    def _get_rows(self):
        wb = load_wb()
        ws = wb["Billing Reconciliation"]
        return sheet_rows(ws)

    def test_acct_001_feb_delta_zero(self):
        """acct_001 Feb: billed=$92, expected=$92 → delta=0.00. No dispute."""
        headers, rows = self._get_rows()
        acct_col  = headers["account_id"] - 1
        month_col = headers["service_month"] - 1
        delta_col = headers["delta_usd"] - 1

        row = next((r for r in rows
                    if str(r[acct_col]).strip() == "acct_001"
                    and str(r[month_col]).strip() == "2026-02"), None)
        assert row is not None, (
            "acct_001 Feb missing from Billing Reconciliation — "
            "Sheet 3 requires one row per account per service month"
        )
        v = float(row[delta_col])
        assert abs(v) <= TOLERANCE, f"acct_001 Feb delta should be 0, got {v}"

    def test_acct_002_jan_delta_overbilled(self):
        """acct_002 Jan: EUR 150 @ 1.0850 = USD 162.75 billed, expected $138 → delta $24.75."""
        headers, rows = self._get_rows()
        acct_col  = headers["account_id"] - 1
        month_col = headers["service_month"] - 1
        delta_col = headers["delta_usd"] - 1
        disc_col  = headers.get("discount_delta_usd", 0) - 1

        row = next((r for r in rows
                    if str(r[acct_col]).strip() == "acct_002"
                    and str(r[month_col]).strip() == "2026-01"), None)
        assert row is not None, "acct_002 Jan missing from Billing Reconciliation"
        v = float(row[delta_col])
        assert abs(v - 24.75) <= TOLERANCE, (
            f"acct_002 Jan delta: expected 24.75 (EUR converted, discount missing), got {v}"
        )
        if disc_col >= 0:
            assert abs(float(row[disc_col]) - 24.75) <= TOLERANCE, (
                f"acct_002 Jan discount_delta_usd: expected 24.75, got {row[disc_col]}"
            )

    def test_acct_002_mar_delta_overbilled(self):
        """acct_002 Mar: EUR 425 @ 1.0790 = USD 458.58 billed, expected $391 → delta $67.58."""
        headers, rows = self._get_rows()
        acct_col  = headers["account_id"] - 1
        month_col = headers["service_month"] - 1
        delta_col = headers["delta_usd"] - 1
        disc_col  = headers.get("discount_delta_usd", 0) - 1

        row = next((r for r in rows
                    if str(r[acct_col]).strip() == "acct_002"
                    and str(r[month_col]).strip() == "2026-03"), None)
        assert row is not None, "acct_002 Mar missing from Billing Reconciliation"
        v = float(row[delta_col])
        assert abs(v - 67.58) <= TOLERANCE, (
            f"acct_002 Mar delta: expected 67.58, got {v}"
        )
        if disc_col >= 0:
            assert abs(float(row[disc_col]) - 67.58) <= TOLERANCE, (
                f"acct_002 Mar discount_delta_usd: expected 67.58, got {row[disc_col]}"
            )

    def test_acct_003_feb_tax_excluded(self):
        """acct_003 Feb: tax=$10.24 billed in overage. delta=$10.24, tax_excluded=$10.24."""
        headers, rows = self._get_rows()
        acct_col   = headers["account_id"] - 1
        month_col  = headers["service_month"] - 1
        delta_col  = headers["delta_usd"] - 1
        tax_col    = headers.get("tax_excluded_usd", 0) - 1
        billed_col = headers.get("billed_overage_usd", 0) - 1

        row = next((r for r in rows
                    if str(r[acct_col]).strip() == "acct_003"
                    and str(r[month_col]).strip() == "2026-02"), None)
        assert row is not None, "acct_003 Feb missing from Billing Reconciliation"
        delta_v = float(row[delta_col])
        assert abs(delta_v - 10.24) <= TOLERANCE, (
            f"acct_003 Feb delta: expected 10.24 (tax-included overbilling), got {delta_v}"
        )
        if tax_col >= 0:
            assert abs(float(row[tax_col]) - 10.24) <= TOLERANCE, (
                f"acct_003 Feb tax_excluded_usd: expected 10.24, got {row[tax_col]}"
            )
        if billed_col >= 0:
            assert abs(float(row[billed_col]) - 74.24) <= TOLERANCE, (
                f"acct_003 Feb billed_overage_usd: expected 74.24, got {row[billed_col]}"
            )

    def test_acct_003_mar_delta_wrong_tier(self):
        """acct_003 Mar: billed $240, expected $230 → delta $10.00."""
        headers, rows = self._get_rows()
        acct_col  = headers["account_id"] - 1
        month_col = headers["service_month"] - 1
        delta_col = headers["delta_usd"] - 1

        row = next((r for r in rows
                    if str(r[acct_col]).strip() == "acct_003"
                    and str(r[month_col]).strip() == "2026-03"), None)
        assert row is not None, "acct_003 Mar missing from Billing Reconciliation"
        v = float(row[delta_col])
        assert abs(v - 10.00) <= TOLERANCE, (
            f"acct_003 Mar delta: expected 10.00 (WRONG_TIER overbilling), got {v}"
        )


# ─────────────────────────────────────────────────────────────────────────────
# SECTION 5: DISPUTE EXCEPTIONS — tests exact string classification
# ─────────────────────────────────────────────────────────────────────────────

class TestDisputeExceptions:
    VALID_TYPES = {"OVERBILLED", "UNDERBILLED", "WRONG_TIER", "TAX_INCLUDED"}
    VALID_SEVERITIES = {"High", "Medium", "Low"}

    def _get_rows(self):
        wb = load_wb()
        ws = wb["Dispute Exceptions"]
        return sheet_rows(ws)

    def test_exception_types_valid(self):
        headers, rows = self._get_rows()
        type_col = headers.get("exception_type", 0) - 1
        if type_col < 0:
            return
        bad = [str(r[type_col]) for r in rows
               if str(r[type_col]).strip() not in self.VALID_TYPES]
        assert not bad, f"Invalid exception_type values: {bad}. Must be one of {self.VALID_TYPES}"

    def test_acct_002_exception_is_overbilled(self):
        headers, rows = self._get_rows()
        acct_col  = headers.get("account_id", 0) - 1
        type_col  = headers.get("exception_type", 0) - 1

        acct_002_exc = [r for r in rows
                        if str(r[acct_col]).strip() == "acct_002"]
        assert acct_002_exc, (
            "No exceptions found for acct_002. Expected OVERBILLED exceptions for Jan and Mar."
        )
        types = [str(r[type_col]).strip() for r in acct_002_exc]
        assert "OVERBILLED" in types, (
            f"acct_002 exception type must be OVERBILLED. Got: {types}"
        )

    def test_acct_002_overbilled_is_medium_severity(self):
        """Discount omission is a calculation error per pricing addendum severity framework."""
        headers, rows = self._get_rows()
        acct_col  = headers.get("account_id", 0) - 1
        sev_col   = headers.get("severity", 0) - 1
        type_col  = headers.get("exception_type", 0) - 1

        overbilled = [r for r in rows
                      if str(r[acct_col]).strip() == "acct_002"
                      and str(r[type_col]).strip() == "OVERBILLED"]
        assert overbilled, "acct_002 OVERBILLED exceptions missing"
        for row in overbilled:
            sev = str(row[sev_col]).strip()
            assert sev == "Medium", (
                f"acct_002 OVERBILLED severity: expected 'Medium' per pricing addendum, got '{sev}'"
            )

    def test_acct_003_exception_is_tax_included(self):
        headers, rows = self._get_rows()
        acct_col  = headers.get("account_id", 0) - 1
        type_col  = headers.get("exception_type", 0) - 1

        acct_003_exc = [r for r in rows
                        if str(r[acct_col]).strip() == "acct_003"]
        assert acct_003_exc, (
            "No exceptions found for acct_003. Expected TAX_INCLUDED exception for Feb."
        )
        types = [str(r[type_col]).strip() for r in acct_003_exc]
        assert "TAX_INCLUDED" in types, (
            f"acct_003 exception type must be TAX_INCLUDED. Got: {types}"
        )

    def test_source_references_present(self):
        headers, rows = self._get_rows()
        src_col = headers.get("source_reference", 0) - 1
        if src_col < 0:
            return
        missing = [str(r[0]) for r in rows
                   if r[src_col] is None or str(r[src_col]).strip() == ""]
        assert not missing, f"Exceptions with no source_reference: {missing}"

    def test_severity_values_valid(self):
        headers, rows = self._get_rows()
        sev_col = headers.get("severity", 0) - 1
        if sev_col < 0:
            return
        bad = [str(r[sev_col]) for r in rows
               if str(r[sev_col]).strip() not in self.VALID_SEVERITIES]
        assert not bad, f"Invalid severity values: {bad}"

    def test_acct_003_is_high_severity(self):
        """acct_003 TAX_INCLUDED is a contractual violation → High severity per pricing addendum."""
        headers, rows = self._get_rows()
        acct_col  = headers.get("account_id", 0) - 1
        sev_col   = headers.get("severity", 0) - 1
        type_col  = headers.get("exception_type", 0) - 1

        tax_exceptions = [r for r in rows
                          if str(r[acct_col]).strip() == "acct_003"
                          and str(r[type_col]).strip() == "TAX_INCLUDED"]
        assert tax_exceptions, "acct_003 TAX_INCLUDED exception missing"
        sev = str(tax_exceptions[0][sev_col]).strip()
        assert sev == "High", (
            f"acct_003 TAX_INCLUDED severity: expected 'High' (contractual violation), got '{sev}'"
        )

    def test_acct_003_mar_exception_is_wrong_tier(self):
        """acct_003 Mar: invoice flat-rated Tier 1 → WRONG_TIER, Medium severity."""
        headers, rows = self._get_rows()
        acct_col  = headers.get("account_id", 0) - 1
        type_col  = headers.get("exception_type", 0) - 1
        sev_col   = headers.get("severity", 0) - 1

        wrong_tier = [r for r in rows
                      if str(r[acct_col]).strip() == "acct_003"
                      and str(r[type_col]).strip() == "WRONG_TIER"]
        assert wrong_tier, (
            "acct_003 WRONG_TIER exception missing for Mar tier arithmetic error"
        )
        sev = str(wrong_tier[0][sev_col]).strip()
        assert sev == "Medium", (
            f"acct_003 WRONG_TIER severity: expected 'Medium' per pricing addendum, got '{sev}'"
        )


# ─────────────────────────────────────────────────────────────────────────────
# SECTION 6: EXECUTIVE SUMMARY — tests cascade from all prior sheets
# ─────────────────────────────────────────────────────────────────────────────

class TestExecutiveSummary:
    def _get_metrics(self):
        wb = load_wb()
        ws = wb["Executive Summary"]
        metrics = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is not None:
                metrics[str(row[0]).strip()] = row[1]
        return metrics

    def test_total_accounts_reviewed(self):
        m = self._get_metrics()
        assert "total_accounts_reviewed" in m
        assert int(m["total_accounts_reviewed"]) == 3, (
            f"total_accounts_reviewed: expected 3, got {m['total_accounts_reviewed']}"
        )

    def test_total_overbilled(self):
        """Total overbilled = 24.75 + 67.58 + 10.24 + 10.00 = 112.57."""
        m = self._get_metrics()
        assert "total_overbilled_usd" in m
        v = float(m["total_overbilled_usd"])
        assert abs(v - 112.57) <= TOLERANCE, (
            f"total_overbilled_usd: expected 112.57, got {v}."
        )

    def test_total_underbilled(self):
        m = self._get_metrics()
        assert "total_underbilled_usd" in m
        v = float(m["total_underbilled_usd"])
        assert abs(v - 0.00) <= TOLERANCE, (
            f"total_underbilled_usd: expected 0.00, got {v}"
        )

    def test_highest_risk_account(self):
        """acct_002 has the largest total delta ($92.33) — should be highest risk."""
        m = self._get_metrics()
        assert "highest_risk_account" in m
        assert str(m["highest_risk_account"]).strip() == "acct_002", (
            f"highest_risk_account: expected 'acct_002' (total delta $92.33), "
            f"got '{m['highest_risk_account']}'"
        )

    def test_recommended_action_immediate(self):
        """Recommended action = 'Immediate Action' because acct_003 TAX_INCLUDED has severity=High."""
        m = self._get_metrics()
        assert "recommended_action" in m
        assert str(m["recommended_action"]).strip() == "Immediate Action", (
            f"recommended_action: expected 'Immediate Action' (acct_003 has High severity exception), "
            f"got '{m['recommended_action']}'."
        )


# ─────────────────────────────────────────────────────────────────────────────
# SECTION 7: OUTPUT FORMAT — numeric literals, not formulas
# ─────────────────────────────────────────────────────────────────────────────

class TestOutputFormat:
    NUMERIC_SUMMARY_METRICS = {
        "total_accounts_reviewed",
        "total_overbilled_usd",
        "total_underbilled_usd",
    }

    def test_summary_metrics_are_literals_not_formulas(self):
        wb = load_wb_formulas()
        ws = wb["Executive Summary"]
        for row in ws.iter_rows(min_row=2):
            metric = row[0].value
            value_cell = row[1]
            if metric is None:
                continue
            metric = str(metric).strip()
            if metric in self.NUMERIC_SUMMARY_METRICS:
                assert value_cell.data_type != "f", (
                    f"Executive Summary '{metric}' must be a computed literal, not a formula"
                )

    def test_reconciliation_deltas_are_literals_not_formulas(self):
        wb = load_wb_formulas()
        ws = wb["Billing Reconciliation"]
        headers, _ = sheet_rows(ws)
        delta_col = headers.get("delta_usd")
        assert delta_col, "delta_usd column missing"
        for row_idx in range(2, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=delta_col)
            if cell.value is None:
                continue
            assert cell.data_type != "f", (
                f"Billing Reconciliation {cell.coordinate} must be a literal, not a formula"
            )
