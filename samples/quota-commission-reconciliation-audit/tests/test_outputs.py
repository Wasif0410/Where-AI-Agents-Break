"""
test_outputs.py — Deterministic pytest verifier for quota-commission-reconciliation-audit.
No LLM-as-judge. All expected values are derived from build_inputs.py canonical data.

Cascading failure design:
  If CSV currency strings are not parsed (Trap A) → base quota wrong → attainment/tier/commission cascade.
  If Excel sheet/formulas are not handled (Trap B) → carryover = 0/NaN → downstream failures.
  If JSONL currency strings are not parsed (Trap C) → bookings sum wrong → attainment/tier/commission cascade.
  If workbook is recreated from scratch (Trap D) → hidden sheets, formulas, validations lost.
"""

from __future__ import annotations

import json
import re
from pathlib import Path

import pytest
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries

OUT_DIR = Path("/root/out")
REPAIRED_XLSX = OUT_DIR / "commission_audit_repaired.xlsx"
SUMMARY_JSON = OUT_DIR / "audit_summary.json"

MONEY_TOL = 0.01
SHARE_TOL = 0.0005

EXPECTED_SHEET_ORDER = [
    "Cover", "Commission Audit", "Summary", "Control", "Validation", "Notes",
]

AUDIT_HEADERS = [
    "rep_id", "base_quota_usd", "carryover_quota_usd", "total_quota_usd",
    "bookings_arr_usd", "attainment_pct", "commission_tier",
    "commission_usd", "draft_commission_usd", "delta_usd", "flag",
    "reviewer_note",
]

EDITABLE_NUMERIC_COLS = (2, 3, 5, 8, 9)
FORMULA_AUDIT_COLS = (4, 6, 7, 10, 11)

TABLE_NAME = "CommissionAuditTable"
EXPECTED_TABLE_RANGE = "A4:L24"

EXPECTED_DEFINED_NAMES = {
    "AuditDataRange": "'Commission Audit'!$A$5:$L$24",
    "TierEntryRange": "'Commission Audit'!$G$5:$G$24",
    "FlagEntryRange": "'Commission Audit'!$K$5:$K$24",
}

EXPECTED_CONDITIONAL_FORMAT_RANGES = {
    "F": "F5:F24",
    "J": "J5:J24",
    "K": "K5:K24",
}

EXPECTED_CONTROL_METADATA = {
    "editable_columns": "B,C,E,H,I",
    "formula_columns": "D,F,G,J,K",
    "preserve_columns": "A,L",
    "first_data_row": 5,
    "last_data_row": 24,
    "active_range": "A4:L24",
    "validation_tier_range": "G5:G24",
    "validation_flag_range": "K5:K24",
    "summary_formula_range": "5:24",
    "table_name": TABLE_NAME,
    "expected_table_range": EXPECTED_TABLE_RANGE,
    "defined_names": "AuditDataRange,TierEntryRange,FlagEntryRange",
    "expected_AuditDataRange": EXPECTED_DEFINED_NAMES["AuditDataRange"],
    "expected_TierEntryRange": EXPECTED_DEFINED_NAMES["TierEntryRange"],
    "expected_FlagEntryRange": EXPECTED_DEFINED_NAMES["FlagEntryRange"],
    "conditional_format_ranges": "F5:F24,J5:J24,K5:K24",
}

HELPER_ROW_FORMULAS = {
    "B25": "=SUM(B5:B24)",
    "C25": "=SUM(C5:C24)",
    "E25": "=SUM(E5:E24)",
    "H25": "=SUM(H5:H24)",
    "J25": "=SUM(J5:J24)",
}

EXPECTED_REVIEWER_NOTES = {
    f"REP-{i:03d}": "Manual review: draft value from payroll export"
    for i in range(1, 21)
}


def audit_row_formulas(row: int) -> dict[str, str]:
    return {
        "D": f"=B{row}+C{row}",
        "F": f"=E{row}/D{row}",
        "G": (
            f'=IF(F{row}<0.8,"Below Threshold",IF(F{row}<1,"Base",'
            f'IF(F{row}<1.2,"Accelerator 1",IF(F{row}<1.5,"Accelerator 2","Accelerator 3"))))'
        ),
        "J": f"=I{row}-H{row}",
        "K": (
            f'=IF(J{row}>0.01,"OVERPAID",IF(J{row}<-0.01,"UNDERPAID","CORRECT"))'
        ),
    }


def tier_and_flag_from_inputs(base, carry, bookings, commission, draft):
    total = base + carry
    att = bookings / total if total else 0.0
    if att < 0.80:
        tier = "Below Threshold"
    elif att < 1.00:
        tier = "Base"
    elif att < 1.20:
        tier = "Accelerator 1"
    elif att < 1.50:
        tier = "Accelerator 2"
    else:
        tier = "Accelerator 3"
    delta = draft - commission
    if delta > 0.01:
        flag = "OVERPAID"
    elif delta < -0.01:
        flag = "UNDERPAID"
    else:
        flag = "CORRECT"
    return tier, flag, delta, total, att

ALLOWED_TIERS = {
    "Below Threshold", "Base", "Accelerator 1",
    "Accelerator 2", "Accelerator 3",
}
ALLOWED_FLAGS = {"OVERPAID", "UNDERPAID", "CORRECT"}

SUMMARY_KEYS = [
    "total_reps_audited", "reps_with_discrepancy", "total_overpaid_usd",
    "total_underpaid_usd", "tier_distribution", "highest_delta_rep",
]

SUMMARY_FORMULAS = {
    "B3": "=COUNTA('Commission Audit'!A5:A24)",
    "B4": "=COUNTIF('Commission Audit'!K5:K24,\"<>CORRECT\")",
    "B5": "=SUMIF('Commission Audit'!K5:K24,\"OVERPAID\",'Commission Audit'!J5:J24)",
    "B6": "=ABS(SUMIF('Commission Audit'!K5:K24,\"UNDERPAID\",'Commission Audit'!J5:J24))",
    "B7": "=IFERROR(INDEX('Commission Audit'!A5:A24,MATCH(MAX(ABS('Commission Audit'!J5:J24)),ABS('Commission Audit'!J5:J24),0)),\"\")",
    "B9": "=COUNTIF('Commission Audit'!G5:G24,\"Below Threshold\")",
    "B10": "=COUNTIF('Commission Audit'!G5:G24,\"Base\")",
    "B11": "=COUNTIF('Commission Audit'!G5:G24,\"Accelerator 1\")",
    "B12": "=COUNTIF('Commission Audit'!G5:G24,\"Accelerator 2\")",
    "B13": "=COUNTIF('Commission Audit'!G5:G24,\"Accelerator 3\")",
}

STALE_SUMMARY_RANGE_FRAGMENTS = (
    "A5:A12", "K5:K12", "J5:J12", "G5:G12",
)


def normalize_summary_formula(formula) -> str:
    """Canonical form: ignore absolute $ markers and sheet-name quote style."""
    if formula is None:
        return ""
    text = str(formula).replace("$", "")
    return re.sub(
        r"['\"]Commission Audit['\"]",
        "'Commission Audit'",
        text,
        flags=re.IGNORECASE,
    )


def assert_summary_formula_correct(cell_ref: str, actual) -> None:
    """Check Summary formulas by structure/ranges, not punctuation-only differences."""
    expected = SUMMARY_FORMULAS[cell_ref]
    assert actual and str(actual).startswith("="), (
        f"Summary!{cell_ref} must remain a formula, got {actual!r}"
    )
    actual_text = str(actual).replace("$", "")
    assert any(tag in actual_text for tag in ("A24", "K24", "J24", "G24")), (
        f"Summary!{cell_ref} must reference row-24 Commission Audit range, got {actual!r}"
    )
    assert not any(stale in actual_text for stale in STALE_SUMMARY_RANGE_FRAGMENTS), (
        f"Summary!{cell_ref} still uses stale 8-rep range: {actual!r}"
    )
    assert normalize_summary_formula(actual) == normalize_summary_formula(expected), (
        f"Summary!{cell_ref} formula structure changed: expected {expected!r}, got {actual!r}"
    )

VALIDATION_TIER_LABELS = [
    "Below Threshold", "Base", "Accelerator 1", "Accelerator 2", "Accelerator 3",
]
VALIDATION_FLAG_LABELS = ["OVERPAID", "UNDERPAID", "CORRECT"]

EXPECTED = {
    "REP-001": {
        "base_quota_usd": 145000.00, "carryover_quota_usd": 15000.00,
        "total_quota_usd": 160000.00, "bookings_arr_usd": 168000.00,
        "attainment_pct": 1.0500, "commission_tier": "Accelerator 1",
        "commission_usd": 16800.00, "draft_commission_usd": 17200.00,
        "delta_usd": 400.00, "flag": "OVERPAID",
    },
    "REP-002": {
        "base_quota_usd": 120000.00, "carryover_quota_usd": 10000.00,
        "total_quota_usd": 130000.00, "bookings_arr_usd": 91000.00,
        "attainment_pct": 0.7000, "commission_tier": "Below Threshold",
        "commission_usd": 0.00, "draft_commission_usd": 0.00,
        "delta_usd": 0.00, "flag": "CORRECT",
    },
    "REP-003": {
        "base_quota_usd": 180000.00, "carryover_quota_usd": 22000.00,
        "total_quota_usd": 202000.00, "bookings_arr_usd": 333300.00,
        "attainment_pct": 1.6500, "commission_tier": "Accelerator 3",
        "commission_usd": 49995.00, "draft_commission_usd": 50500.00,
        "delta_usd": 505.00, "flag": "OVERPAID",
    },
    "REP-004": {
        "base_quota_usd": 100000.00, "carryover_quota_usd": 5000.00,
        "total_quota_usd": 105000.00, "bookings_arr_usd": 94500.00,
        "attainment_pct": 0.9000, "commission_tier": "Base",
        "commission_usd": 7560.00, "draft_commission_usd": 7560.00,
        "delta_usd": 0.00, "flag": "CORRECT",
    },
    "REP-005": {
        "base_quota_usd": 200000.00, "carryover_quota_usd": 30000.00,
        "total_quota_usd": 230000.00, "bookings_arr_usd": 310500.00,
        "attainment_pct": 1.3500, "commission_tier": "Accelerator 2",
        "commission_usd": 37260.00, "draft_commission_usd": 37260.00,
        "delta_usd": 0.00, "flag": "CORRECT",
    },
    "REP-006": {
        "base_quota_usd": 40000.00, "carryover_quota_usd": 4000.00,
        "total_quota_usd": 44000.00, "bookings_arr_usd": 33750.00,
        "attainment_pct": 0.7670, "commission_tier": "Below Threshold",
        "commission_usd": 0.00, "draft_commission_usd": 0.00,
        "delta_usd": 0.00, "flag": "CORRECT",
    },
    "REP-007": {
        "base_quota_usd": 160000.00, "carryover_quota_usd": 16000.00,
        "total_quota_usd": 176000.00, "bookings_arr_usd": 193600.00,
        "attainment_pct": 1.1000, "commission_tier": "Accelerator 1",
        "commission_usd": 19360.00, "draft_commission_usd": 18800.00,
        "delta_usd": -560.00, "flag": "UNDERPAID",
    },
    "REP-008": {
        "base_quota_usd": 95000.00, "carryover_quota_usd": 0.00,
        "total_quota_usd": 95000.00, "bookings_arr_usd": 83600.00,
        "attainment_pct": 0.8800, "commission_tier": "Base",
        "commission_usd": 6688.00, "draft_commission_usd": 6688.00,
        "delta_usd": 0.00, "flag": "CORRECT",
    },
    "REP-009": {
        "base_quota_usd": 55000.00, "carryover_quota_usd": 4000.00,
        "total_quota_usd": 59000.00, "bookings_arr_usd": 50200.00,
        "attainment_pct": 0.8508, "commission_tier": "Base",
        "commission_usd": 4016.00, "draft_commission_usd": 4016.00,
        "delta_usd": 0.00, "flag": "CORRECT",
    },
    "REP-010": {
        "base_quota_usd": 75000.00, "carryover_quota_usd": 7000.00,
        "total_quota_usd": 82000.00, "bookings_arr_usd": 57500.00,
        "attainment_pct": 0.7012, "commission_tier": "Below Threshold",
        "commission_usd": 0.00, "draft_commission_usd": 0.00,
        "delta_usd": 0.00, "flag": "CORRECT",
    },
    "REP-011": {
        "base_quota_usd": 110000.00, "carryover_quota_usd": 12000.00,
        "total_quota_usd": 122000.00, "bookings_arr_usd": 128100.00,
        "attainment_pct": 1.0500, "commission_tier": "Accelerator 1",
        "commission_usd": 12810.00, "draft_commission_usd": 12810.00,
        "delta_usd": 0.00, "flag": "CORRECT",
    },
    "REP-012": {
        "base_quota_usd": 65000.00, "carryover_quota_usd": 5000.00,
        "total_quota_usd": 70000.00, "bookings_arr_usd": 103500.00,
        "attainment_pct": 1.4786, "commission_tier": "Accelerator 2",
        "commission_usd": 12420.00, "draft_commission_usd": 11500.00,
        "delta_usd": -920.00, "flag": "UNDERPAID",
    },
    "REP-013": {
        "base_quota_usd": 220000.00, "carryover_quota_usd": 33000.00,
        "total_quota_usd": 253000.00, "bookings_arr_usd": 404800.00,
        "attainment_pct": 1.6000, "commission_tier": "Accelerator 3",
        "commission_usd": 60720.00, "draft_commission_usd": 60720.00,
        "delta_usd": 0.00, "flag": "CORRECT",
    },
    "REP-014": {
        "base_quota_usd": 85000.00, "carryover_quota_usd": 7000.00,
        "total_quota_usd": 92000.00, "bookings_arr_usd": 84640.00,
        "attainment_pct": 0.9200, "commission_tier": "Base",
        "commission_usd": 6771.20, "draft_commission_usd": 6771.20,
        "delta_usd": 0.00, "flag": "CORRECT",
    },
    "REP-015": {
        "base_quota_usd": 50000.00, "carryover_quota_usd": 5000.00,
        "total_quota_usd": 55000.00, "bookings_arr_usd": 39750.00,
        "attainment_pct": 0.7227, "commission_tier": "Below Threshold",
        "commission_usd": 0.00, "draft_commission_usd": 0.00,
        "delta_usd": 0.00, "flag": "CORRECT",
    },
    "REP-016": {
        "base_quota_usd": 135000.00, "carryover_quota_usd": 16000.00,
        "total_quota_usd": 151000.00, "bookings_arr_usd": 173650.00,
        "attainment_pct": 1.1500, "commission_tier": "Accelerator 1",
        "commission_usd": 17365.00, "draft_commission_usd": 17365.00,
        "delta_usd": 0.00, "flag": "CORRECT",
    },
    "REP-017": {
        "base_quota_usd": 70000.00, "carryover_quota_usd": 7000.00,
        "total_quota_usd": 77000.00, "bookings_arr_usd": 107800.00,
        "attainment_pct": 1.4000, "commission_tier": "Accelerator 2",
        "commission_usd": 12936.00, "draft_commission_usd": 12936.00,
        "delta_usd": 0.00, "flag": "CORRECT",
    },
    "REP-018": {
        "base_quota_usd": 48000.00, "carryover_quota_usd": 3000.00,
        "total_quota_usd": 51000.00, "bookings_arr_usd": 44880.00,
        "attainment_pct": 0.8800, "commission_tier": "Base",
        "commission_usd": 3590.40, "draft_commission_usd": 3590.40,
        "delta_usd": 0.00, "flag": "CORRECT",
    },
    "REP-019": {
        "base_quota_usd": 175000.00, "carryover_quota_usd": 23000.00,
        "total_quota_usd": 198000.00, "bookings_arr_usd": 306900.00,
        "attainment_pct": 1.5500, "commission_tier": "Accelerator 3",
        "commission_usd": 46035.00, "draft_commission_usd": 46035.00,
        "delta_usd": 0.00, "flag": "CORRECT",
    },
    "REP-020": {
        "base_quota_usd": 60000.00, "carryover_quota_usd": 0.00,
        "total_quota_usd": 60000.00, "bookings_arr_usd": 36000.00,
        "attainment_pct": 0.6000, "commission_tier": "Below Threshold",
        "commission_usd": 0.00, "draft_commission_usd": 0.00,
        "delta_usd": 0.00, "flag": "CORRECT",
    },
}

EXPECTED_SUMMARY = {
    "total_reps_audited": 20,
    "reps_with_discrepancy": 4,
    "total_overpaid_usd": 905.00,
    "total_underpaid_usd": 1480.00,
    "tier_distribution": {
        "Below Threshold": 5,
        "Base": 5,
        "Accelerator 1": 4,
        "Accelerator 2": 3,
        "Accelerator 3": 3,
    },
    "highest_delta_rep": "REP-012",
}

STALE_REP001_CARRYOVER = 0


def load_workbook_repaired():
    assert REPAIRED_XLSX.exists(), f"Missing: {REPAIRED_XLSX}"
    return load_workbook(REPAIRED_XLSX, data_only=False)


def load_summary():
    assert SUMMARY_JSON.exists(), f"Missing: {SUMMARY_JSON}"
    with open(SUMMARY_JSON, encoding="utf-8") as f:
        return json.load(f)


def cf_range_includes_row(ws, min_col: int, max_col: int, row: int) -> bool:
    """Return True if any conditional-formatting rule covers the column band through row."""
    for cf in ws.conditional_formatting:
        for cr in str(cf.sqref).split():
            c_min, r_min, c_max, r_max = range_boundaries(cr)
            if r_min <= row <= r_max and c_min <= min_col and max_col <= c_max:
                return True
    return False


def defined_name_destination(wb, name: str) -> str | None:
    dn = wb.defined_names.get(name)
    if dn is None:
        return None
    return getattr(dn, "attr_text", None) or getattr(dn, "value", None)


def audit_rows_from_workbook(wb):
    """Read editable cells and derive formula-column values from inputs."""
    ws = wb["Commission Audit"]
    rows = {}
    for row_idx in range(5, 25):
        rep_id = ws.cell(row=row_idx, column=1).value
        base = float(ws.cell(row=row_idx, column=2).value)
        carry = float(ws.cell(row=row_idx, column=3).value)
        bookings = float(ws.cell(row=row_idx, column=5).value)
        commission = float(ws.cell(row=row_idx, column=8).value)
        draft = float(ws.cell(row=row_idx, column=9).value)
        tier, flag, delta, total, att = tier_and_flag_from_inputs(
            base, carry, bookings, commission, draft
        )
        rows[rep_id] = {
            "base_quota_usd": base,
            "carryover_quota_usd": carry,
            "total_quota_usd": total,
            "bookings_arr_usd": bookings,
            "attainment_pct": att,
            "commission_tier": tier,
            "commission_usd": commission,
            "draft_commission_usd": draft,
            "delta_usd": delta,
            "flag": flag,
        }
    return rows


class TestOutputFilesExist:
    def test_repaired_workbook_exists(self):
        assert REPAIRED_XLSX.exists(), (
            "commission_audit_repaired.xlsx not found in /root/out"
        )

    def test_audit_summary_json_exists(self):
        assert SUMMARY_JSON.exists(), "audit_summary.json not found in /root/out"


class TestWorkbookStructure:
    def test_sheet_names_and_order(self):
        wb = load_workbook_repaired()
        assert wb.sheetnames == EXPECTED_SHEET_ORDER, (
            f"Sheet order must be {EXPECTED_SHEET_ORDER}, got {wb.sheetnames}"
        )

    def test_control_sheet_hidden(self):
        wb = load_workbook_repaired()
        assert "Control" in wb.sheetnames
        assert wb["Control"].sheet_state == "hidden"

    def test_validation_sheet_hidden(self):
        wb = load_workbook_repaired()
        assert "Validation" in wb.sheetnames
        assert wb["Validation"].sheet_state == "hidden"

    def test_notes_sheet_exists(self):
        wb = load_workbook_repaired()
        assert "Notes" in wb.sheetnames


class TestCommissionAuditLayout:
    def test_headers(self):
        wb = load_workbook_repaired()
        ws = wb["Commission Audit"]
        headers = [ws.cell(row=4, column=c).value for c in range(1, 13)]
        assert headers == AUDIT_HEADERS

    def test_rep_ids(self):
        wb = load_workbook_repaired()
        ws = wb["Commission Audit"]
        ids = [ws.cell(row=r, column=1).value for r in range(5, 25)]
        expected_ids = [f"REP-{i:03d}" for i in range(1, 21)]
        assert ids == expected_ids

    def test_frozen_panes(self):
        wb = load_workbook_repaired()
        assert wb["Commission Audit"].freeze_panes == "A5"

    def test_auto_filter(self):
        wb = load_workbook_repaired()
        assert wb["Commission Audit"].auto_filter.ref == "A4:L24"


class TestWorkbookPreservation:
    def test_summary_formulas_preserved(self):
        wb = load_workbook_repaired()
        ws = wb["Summary"]
        for cell_ref in SUMMARY_FORMULAS:
            assert_summary_formula_correct(cell_ref, ws[cell_ref].value)

    def test_control_expected_row_count(self):
        wb = load_workbook_repaired()
        assert wb["Control"]["B2"].value == 20

    def test_control_metadata_preserved(self):
        wb = load_workbook_repaired()
        ws = wb["Control"]
        for key, expected in EXPECTED_CONTROL_METADATA.items():
            found = None
            for row in range(1, 35):
                if ws.cell(row=row, column=1).value == key:
                    found = ws.cell(row=row, column=2).value
                    break
            assert found == expected, (
                f"Control!{key}: expected {expected!r}, got {found!r}"
            )

    def test_validation_tier_labels(self):
        wb = load_workbook_repaired()
        ws = wb["Validation"]
        labels = [ws.cell(row=r, column=1).value for r in range(2, 7)]
        assert labels == VALIDATION_TIER_LABELS

    def test_validation_flag_labels(self):
        wb = load_workbook_repaired()
        ws = wb["Validation"]
        labels = [ws.cell(row=r, column=3).value for r in range(2, 5)]
        assert labels == VALIDATION_FLAG_LABELS

    def test_tier_data_validation(self):
        wb = load_workbook_repaired()
        ws = wb["Commission Audit"]
        assert ws.data_validations.dataValidation, "No data validations on Commission Audit"
        refs = [str(dv.sqref) for dv in ws.data_validations.dataValidation if "G" in str(dv.sqref)]
        assert refs, "Tier column data validation missing"
        assert any("G24" in ref or "G5:G24" in ref for ref in refs), (
            f"Tier validation must cover G5:G24, got {refs}"
        )

    def test_flag_data_validation(self):
        wb = load_workbook_repaired()
        ws = wb["Commission Audit"]
        refs = [str(dv.sqref) for dv in ws.data_validations.dataValidation if "K" in str(dv.sqref)]
        assert refs, "Flag column data validation missing"
        assert any("K24" in ref or "K5:K24" in ref for ref in refs), (
            f"Flag validation must cover K5:K24, got {refs}"
        )

    def test_header_row_styled(self):
        wb = load_workbook_repaired()
        header = wb["Commission Audit"].cell(row=4, column=1)
        assert header.font.bold, "Header row styling was overwritten"
        assert header.fill.fill_type == "solid", "Header fill was overwritten"


class TestEditableCellsAreLiterals:
    def test_editable_numeric_cells_not_formulas(self):
        wb = load_workbook_repaired()
        ws = wb["Commission Audit"]
        for row in range(5, 25):
            for col in EDITABLE_NUMERIC_COLS:
                cell = ws.cell(row=row, column=col)
                assert cell.data_type != "f", (
                    f"Editable cell {cell.coordinate} must be a literal value, not a formula"
                )

    def test_editable_numeric_cells_are_numbers(self):
        wb = load_workbook_repaired()
        ws = wb["Commission Audit"]
        for row in range(5, 25):
            for col in EDITABLE_NUMERIC_COLS:
                val = ws.cell(row=row, column=col).value
                assert isinstance(val, (int, float)), (
                    f"Editable cell row {row} col {col} must be numeric, got {type(val).__name__}: {val!r}"
                )


class TestAuditFormulaCells:
    def test_total_quota_formulas_preserved(self):
        wb = load_workbook_repaired()
        ws = wb["Commission Audit"]
        for row in range(5, 25):
            cell = ws.cell(row=row, column=4)
            assert cell.data_type == "f", f"D{row} must remain a formula"
            assert cell.value == audit_row_formulas(row)["D"]

    def test_attainment_formulas_preserved(self):
        wb = load_workbook_repaired()
        ws = wb["Commission Audit"]
        for row in range(5, 25):
            cell = ws.cell(row=row, column=6)
            assert cell.data_type == "f", f"F{row} must remain a formula"
            assert cell.value == audit_row_formulas(row)["F"]

    def test_delta_formulas_preserved(self):
        wb = load_workbook_repaired()
        ws = wb["Commission Audit"]
        for row in range(5, 25):
            cell = ws.cell(row=row, column=10)
            assert cell.data_type == "f", f"J{row} must remain a formula"
            assert cell.value == audit_row_formulas(row)["J"]

    def test_tier_formulas_preserved(self):
        wb = load_workbook_repaired()
        ws = wb["Commission Audit"]
        for row in range(5, 25):
            cell = ws.cell(row=row, column=7)
            assert cell.data_type == "f", f"G{row} must remain a formula"
            assert cell.value == audit_row_formulas(row)["G"]

    def test_flag_formulas_preserved(self):
        wb = load_workbook_repaired()
        ws = wb["Commission Audit"]
        for row in range(5, 25):
            cell = ws.cell(row=row, column=11)
            assert cell.data_type == "f", f"K{row} must remain a formula"
            assert cell.value == audit_row_formulas(row)["K"]

    def test_hidden_helper_row_preserved(self):
        wb = load_workbook_repaired()
        ws = wb["Commission Audit"]
        assert ws.row_dimensions[25].hidden is True
        for cell_ref, expected_formula in HELPER_ROW_FORMULAS.items():
            cell = ws[cell_ref]
            assert cell.data_type == "f", f"{cell_ref} must remain a formula"
            assert cell.value == expected_formula

    def test_reviewer_notes_preserved(self):
        wb = load_workbook_repaired()
        ws = wb["Commission Audit"]
        for row in range(5, 25):
            rep_id = ws.cell(row=row, column=1).value
            actual = ws.cell(row=row, column=12).value
            expected = EXPECTED_REVIEWER_NOTES[rep_id]
            assert actual == expected, (
                f"{rep_id} reviewer_note changed: expected {expected!r}, got {actual!r}"
            )


class TestPerRepValues:
    @pytest.fixture
    def audit_rows(self):
        return audit_rows_from_workbook(load_workbook_repaired())

    def test_all_reps_present(self, audit_rows):
        assert set(audit_rows) == set(EXPECTED)

    def test_base_quota_per_rep(self, audit_rows):
        for rep_id, exp in EXPECTED.items():
            actual = float(audit_rows[rep_id]["base_quota_usd"])
            assert abs(actual - exp["base_quota_usd"]) <= MONEY_TOL, (
                f"{rep_id} base_quota_usd: expected {exp['base_quota_usd']}, got {actual}"
            )

    def test_carryover_nonzero(self, audit_rows):
        total = sum(float(r["carryover_quota_usd"]) for r in audit_rows.values())
        assert total > 0, (
            "All carryover values are zero. Check quota_carryover.xlsx — "
            "carryover amounts must be read from the workbook, not assumed zero."
        )

    def test_carryover_per_rep(self, audit_rows):
        for rep_id, exp in EXPECTED.items():
            actual = float(audit_rows[rep_id]["carryover_quota_usd"])
            assert abs(actual - exp["carryover_quota_usd"]) <= MONEY_TOL, (
                f"{rep_id} carryover_quota_usd: expected {exp['carryover_quota_usd']}, got {actual}"
            )

    def test_total_quota_equals_sum(self, audit_rows):
        for rep_id, row in audit_rows.items():
            calc = float(row["base_quota_usd"]) + float(row["carryover_quota_usd"])
            actual = float(row["total_quota_usd"])
            assert abs(calc - actual) <= MONEY_TOL, (
                f"{rep_id} total_quota_usd must equal base + carryover"
            )

    def test_attainment_per_rep(self, audit_rows):
        for rep_id, exp in EXPECTED.items():
            actual = float(audit_rows[rep_id]["attainment_pct"])
            assert abs(actual - exp["attainment_pct"]) <= SHARE_TOL, (
                f"{rep_id} attainment_pct: expected {exp['attainment_pct']}, got {actual}"
            )

    def test_not_all_below_threshold(self, audit_rows):
        tiers = [r["commission_tier"] for r in audit_rows.values()]
        assert not all(t == "Below Threshold" for t in tiers), (
            "All reps are Below Threshold — base or carryover quotas were likely parsed incorrectly."
        )

    def test_commission_tier_per_rep(self, audit_rows):
        for rep_id, exp in EXPECTED.items():
            actual = audit_rows[rep_id]["commission_tier"]
            assert actual == exp["commission_tier"], (
                f"{rep_id} commission_tier: expected {exp['commission_tier']}, got {actual}"
            )
            assert actual in ALLOWED_TIERS

    def test_commission_usd_per_rep(self, audit_rows):
        for rep_id, exp in EXPECTED.items():
            actual = float(audit_rows[rep_id]["commission_usd"])
            assert abs(actual - exp["commission_usd"]) <= MONEY_TOL, (
                f"{rep_id} commission_usd: expected {exp['commission_usd']}, got {actual}"
            )

    def test_delta_usd_per_rep(self, audit_rows):
        for rep_id, exp in EXPECTED.items():
            actual = float(audit_rows[rep_id]["delta_usd"])
            assert abs(actual - exp["delta_usd"]) <= MONEY_TOL, (
                f"{rep_id} delta_usd: expected {exp['delta_usd']}, got {actual}"
            )

    def test_flag_per_rep(self, audit_rows):
        for rep_id, exp in EXPECTED.items():
            actual = audit_rows[rep_id]["flag"]
            assert actual == exp["flag"], (
                f"{rep_id} flag: expected {exp['flag']}, got {actual}"
            )
            assert actual in ALLOWED_FLAGS


class TestAuditSummary:
    def test_required_keys(self):
        summary = load_summary()
        for key in SUMMARY_KEYS:
            assert key in summary, f"audit_summary.json missing key: '{key}'"

    def test_total_reps_audited(self):
        assert load_summary()["total_reps_audited"] == 20

    def test_reps_with_discrepancy(self):
        assert load_summary()["reps_with_discrepancy"] == 4

    def test_total_overpaid_usd(self):
        actual = float(load_summary()["total_overpaid_usd"])
        expected = EXPECTED_SUMMARY["total_overpaid_usd"]
        assert abs(actual - expected) <= MONEY_TOL

    def test_total_underpaid_usd(self):
        actual = float(load_summary()["total_underpaid_usd"])
        expected = EXPECTED_SUMMARY["total_underpaid_usd"]
        assert abs(actual - expected) <= MONEY_TOL

    def test_tier_distribution(self):
        summary = load_summary()
        dist = summary["tier_distribution"]
        for tier, count in EXPECTED_SUMMARY["tier_distribution"].items():
            assert tier in dist, f"tier_distribution missing tier '{tier}'"
            assert dist[tier] == count, (
                f"tier_distribution['{tier}']: expected {count}, got {dist[tier]}"
            )

    def test_highest_delta_rep(self):
        assert load_summary()["highest_delta_rep"] == "REP-012"

    def test_summary_matches_workbook(self):
        wb = load_workbook_repaired()
        rows = audit_rows_from_workbook(wb)
        summary = load_summary()

        discrepancy = sum(1 for r in rows.values() if r["flag"] != "CORRECT")
        assert summary["reps_with_discrepancy"] == discrepancy

        overpaid = round(
            sum(float(r["delta_usd"]) for r in rows.values() if r["flag"] == "OVERPAID"),
            2,
        )
        underpaid = round(
            sum(abs(float(r["delta_usd"])) for r in rows.values() if r["flag"] == "UNDERPAID"),
            2,
        )
        assert abs(float(summary["total_overpaid_usd"]) - overpaid) <= MONEY_TOL
        assert abs(float(summary["total_underpaid_usd"]) - underpaid) <= MONEY_TOL

        ranked = sorted(
            rows.items(),
            key=lambda item: (-abs(float(item[1]["delta_usd"])), item[0]),
        )
        assert summary["highest_delta_rep"] == ranked[0][0]

        for tier in ALLOWED_TIERS:
            wb_count = sum(1 for r in rows.values() if r["commission_tier"] == tier)
            assert summary["tier_distribution"][tier] == wb_count


class TestWorkbookArtifactRepair:
    def test_commission_audit_table_range(self):
        wb = load_workbook_repaired()
        ws = wb["Commission Audit"]
        assert TABLE_NAME in ws.tables, (
            f"Excel table {TABLE_NAME!r} missing from Commission Audit sheet"
        )
        assert ws.tables[TABLE_NAME].ref == EXPECTED_TABLE_RANGE, (
            f"{TABLE_NAME}.ref must be {EXPECTED_TABLE_RANGE!r}, "
            f"got {ws.tables[TABLE_NAME].ref!r}"
        )

    def test_defined_names_point_to_full_range(self):
        wb = load_workbook_repaired()
        for name, expected in EXPECTED_DEFINED_NAMES.items():
            actual = defined_name_destination(wb, name)
            assert actual is not None, f"Workbook defined name {name!r} is missing"
            assert actual == expected, (
                f"Defined name {name!r}: expected {expected!r}, got {actual!r}"
            )

    def test_conditional_formatting_covers_row_24(self):
        wb = load_workbook_repaired()
        ws = wb["Commission Audit"]
        assert cf_range_includes_row(ws, 6, 6, 24), (
            "Attainment conditional formatting must cover F5:F24"
        )
        assert cf_range_includes_row(ws, 10, 10, 24), (
            "Delta conditional formatting must cover J5:J24"
        )
        assert cf_range_includes_row(ws, 11, 11, 24), (
            "Flag conditional formatting must cover K5:K24"
        )
        for col_letter, expected_range in EXPECTED_CONDITIONAL_FORMAT_RANGES.items():
            found = False
            for cf in ws.conditional_formatting:
                if expected_range.replace(":", "") in str(cf.sqref).replace(":", ""):
                    found = True
                    break
                if expected_range in str(cf.sqref):
                    found = True
                    break
            assert found or cf_range_includes_row(
                ws,
                range_boundaries(expected_range)[0],
                range_boundaries(expected_range)[2],
                24,
            ), (
                f"Conditional formatting for column {col_letter} must extend to row 24"
            )

    def test_formula_columns_extended_through_row_24(self):
        wb = load_workbook_repaired()
        ws = wb["Commission Audit"]
        for row in range(13, 25):
            for col in FORMULA_AUDIT_COLS:
                cell = ws.cell(row=row, column=col)
                assert cell.data_type == "f", (
                    f"{cell.coordinate} must be a formula after repair, not a stale literal"
                )
                assert cell.value == audit_row_formulas(row)[
                    {4: "D", 6: "F", 7: "G", 10: "J", 11: "K"}[col]
                ], (
                    f"{cell.coordinate} formula must be row-relative for row {row}"
                )

    def test_rows_13_24_formulas_reference_own_row(self):
        wb = load_workbook_repaired()
        ws = wb["Commission Audit"]
        for row in range(13, 25):
            d_formula = ws.cell(row=row, column=4).value
            assert f"B{row}" in d_formula and f"C{row}" in d_formula, (
                f"D{row} must reference B{row} and C{row}, got {d_formula!r}"
            )
            f_formula = ws.cell(row=row, column=6).value
            assert f"E{row}" in f_formula and f"D{row}" in f_formula, (
                f"F{row} must reference E{row} and D{row}, got {f_formula!r}"
            )


class TestStaleRangeRepair:
    def test_summary_formulas_use_full_rep_range(self):
        wb = load_workbook_repaired()
        ws = wb["Summary"]
        for cell_ref in SUMMARY_FORMULAS:
            assert_summary_formula_correct(cell_ref, ws[cell_ref].value)

    def test_autofilter_covers_full_range(self):
        wb = load_workbook_repaired()
        assert wb["Commission Audit"].auto_filter.ref == "A4:L24"

    def test_helper_formulas_use_row_24(self):
        wb = load_workbook_repaired()
        ws = wb["Commission Audit"]
        for cell_ref, formula in HELPER_ROW_FORMULAS.items():
            assert "B24" in formula or "C24" in formula or "E24" in formula or "H24" in formula or "J24" in formula
            assert "B12" not in formula and "J12" not in formula
            assert ws[cell_ref].value == formula


class TestAntiShortcut:
    def test_hidden_sheets_not_dropped(self):
        wb = load_workbook_repaired()
        assert wb["Control"].sheet_state == "hidden"
        assert wb["Validation"].sheet_state == "hidden"

    def test_summary_not_overwritten_with_values(self):
        wb = load_workbook_repaired()
        ws = wb["Summary"]
        for cell_ref in ("B3", "B4", "B5", "B6", "B9", "B10", "B11", "B12", "B13"):
            val = ws[cell_ref].value
            assert isinstance(val, str) and val.startswith("="), (
                f"Summary!{cell_ref} must remain a formula, got {val!r}"
            )

    def test_tier_and_flag_not_written_as_literals(self):
        """Gemini habit: write computed tier/flag into G and K."""
        wb = load_workbook_repaired()
        ws = wb["Commission Audit"]
        for row in range(5, 25):
            assert ws.cell(row=row, column=7).data_type == "f"
            assert ws.cell(row=row, column=11).data_type == "f"

    def test_audit_formula_columns_not_overwritten(self):
        """Naive write-all-columns repair overwrites D/F/G/J/K formulas."""
        wb = load_workbook_repaired()
        ws = wb["Commission Audit"]
        for row in range(5, 25):
            for col in FORMULA_AUDIT_COLS:
                cell = ws.cell(row=row, column=col)
                assert cell.data_type == "f", (
                    f"{cell.coordinate} must remain a formula, not a literal value"
                )

    def test_stale_draft_rep001_repaired(self):
        """REP-001 draft had carryover 0 and flag CORRECT; repair must fix both."""
        rows = audit_rows_from_workbook(load_workbook_repaired())
        assert float(rows["REP-001"]["carryover_quota_usd"]) != STALE_REP001_CARRYOVER
        assert rows["REP-001"]["flag"] == "OVERPAID"
        assert abs(float(rows["REP-001"]["carryover_quota_usd"]) - 15000.00) <= MONEY_TOL

    def test_rep012_carryover_uses_max_formula(self):
        rows = audit_rows_from_workbook(load_workbook_repaired())
        carry = float(rows["REP-012"]["carryover_quota_usd"])
        assert abs(carry - 5000.00) <= MONEY_TOL, (
            f"REP-012 carryover_quota_usd is {carry} — expected 5000.00. "
            "The carryover cell uses a MAX formula; arithmetic-only handling gives 4000."
        )

    def test_rep012_wrong_carryover_breaks_tier_and_commission(self):
        """ROUND-only carryover for REP-012 lands in Accelerator 3, not Accelerator 2."""
        rows = audit_rows_from_workbook(load_workbook_repaired())
        assert rows["REP-012"]["commission_tier"] == "Accelerator 2"
        assert abs(float(rows["REP-012"]["commission_usd"]) - 12420.00) <= MONEY_TOL

    def test_rep015_carryover_uses_max_formula(self):
        rows = audit_rows_from_workbook(load_workbook_repaired())
        carry = float(rows["REP-015"]["carryover_quota_usd"])
        assert abs(carry - 5000.00) <= MONEY_TOL, (
            f"REP-015 carryover_quota_usd is {carry} — expected 5000.00. "
            "The carryover cell uses a MAX formula; arithmetic-only handling gives 3000."
        )

    def test_rep008_zero_carryover_preserved(self):
        rows = audit_rows_from_workbook(load_workbook_repaired())
        carry = float(rows["REP-008"]["carryover_quota_usd"])
        assert abs(carry) <= MONEY_TOL

    def test_rep020_zero_carryover_preserved(self):
        rows = audit_rows_from_workbook(load_workbook_repaired())
        carry = float(rows["REP-020"]["carryover_quota_usd"])
        assert abs(carry) <= MONEY_TOL
