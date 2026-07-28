#!/usr/bin/env python3
"""
inspect_workbook.py — diagnostic tool for exploring Excel workbook structure.

Usage:
    python3 scripts/inspect_workbook.py <workbook.xlsx>
    python3 scripts/inspect_workbook.py <workbook.xlsx> <sheet_name>

With no sheet name: lists all sheet names in the workbook.
With a sheet name: prints the header row and first 10 data rows, showing
raw cell values (including formula strings) so you can see what is actually
stored in each cell before deciding how to load the data.
"""
from __future__ import annotations

import sys
from pathlib import Path

import openpyxl


def list_sheets(path: Path) -> None:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=False)
    print(f"Sheets in {path.name}:")
    for name in wb.sheetnames:
        print(f"  {name!r}")


def preview_sheet(path: Path, sheet_name: str, max_rows: int = 10) -> None:
    wb = openpyxl.load_workbook(path, data_only=False)
    if sheet_name not in wb.sheetnames:
        raise SystemExit(
            f"Sheet {sheet_name!r} not found. Available: {wb.sheetnames}"
        )
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=False))
    if not rows:
        print("(empty sheet)")
        return

    print(f"Sheet: {sheet_name!r}  ({ws.max_row} rows x {ws.max_column} cols)")
    print()
    for i, row in enumerate(rows[: max_rows + 1]):
        label = "header" if i == 0 else f"row {i:>3}"
        values = [repr(cell.value) for cell in row]
        print(f"  {label}: {', '.join(values)}")
    if ws.max_row > max_rows + 1:
        print(f"  ... ({ws.max_row - max_rows - 1} more rows)")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        raise SystemExit(__doc__)
    path = Path(sys.argv[1])
    if len(sys.argv) == 2:
        list_sheets(path)
    else:
        preview_sheet(path, sys.argv[2])
